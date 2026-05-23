from __future__ import annotations

import os
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from app import extract_validated_candidates, rejection_reason_for_name


XLSX_PATH = Path(r"C:\Users\HP\Downloads\pep_screening_template.xlsx")
PDF_PATH = Path(r"C:\Users\HP\Downloads\Database.pdf")


def main() -> None:
    os.environ["OPENAI_API_KEY"] = ""
    backup_path = XLSX_PATH.with_name(f"{XLSX_PATH.stem}_backup_{datetime.now():%Y%m%d_%H%M%S}{XLSX_PATH.suffix}")
    shutil.copy2(XLSX_PATH, backup_path)

    reject_terms = {
        "act",
        "authority",
        "business",
        "chamber",
        "clinic",
        "clinics",
        "commission",
        "cooperation",
        "correctional",
        "council",
        "culture",
        "development",
        "forces",
        "fund",
        "human",
        "innovation",
        "international",
        "mission",
        "page",
        "pageant",
        "planning",
        "prison",
        "service",
        "services",
        "settlement",
        "settlements",
        "society",
        "station",
        "technology",
    }
    records: dict[str, tuple[str, str, str, int]] = {}
    reader = PdfReader(str(PDF_PATH))

    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if not text.strip():
            continue
        candidates, _metrics = extract_validated_candidates(
            text,
            source_name="Database.pdf",
            source_url=f"Database.pdf page {index}",
            source_jurisdiction="Botswana",
            log_rejections=False,
        )
        for candidate in candidates:
            name = str(candidate.get("name", "")).strip()
            score = int(candidate.get("confidence_score", 0))
            if score < 95:
                continue
            if rejection_reason_for_name(name):
                continue
            parts = name.lower().split()
            if len(parts) < 2 or len(parts) > 5:
                continue
            if any(term in parts for term in reject_terms):
                continue
            note = (
                f"Category: {candidate.get('category', '')}; "
                f"Position: {candidate.get('position', '')}; "
                f"Jurisdiction: {candidate.get('jurisdiction', 'Botswana')}; "
                f"Confidence: {score}; "
                f"Evidence: {str(candidate.get('snippet', ''))[:600]}"
            )
            key = name.lower()
            row = (name, f"Database.pdf page {index}", note, score)
            if key not in records or score > records[key][3]:
                records[key] = row

    workbook = load_workbook(XLSX_PATH)
    sheet = workbook.active
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(["Name", "Reference", "Notes"])
    for name, reference, notes, _score in sorted(records.values(), key=lambda item: item[0]):
        sheet.append([name, reference, notes])
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 120
    workbook.save(XLSX_PATH)

    print(f"Backup: {backup_path}")
    print(f"Rows written: {len(records)}")
    print(f"Updated: {XLSX_PATH}")


if __name__ == "__main__":
    main()
