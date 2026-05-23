from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app import rejection_reason_for_name


WORKBOOK_PATH = Path(r"C:\Users\HP\Downloads\pep_pip_records_export_20260516.xlsx")


CONFIRMING_ROLE_TERMS = {
    "president",
    "vice president",
    "prime minister",
    "minister",
    "assistant minister",
    "permanent secretary",
    "member of parliament",
    " mp ",
    "speaker",
    "deputy speaker",
    "mayor",
    "councillor",
    "kgosi",
    "chief justice",
    "justice",
    "judge",
    "commissioner",
    "ambassador",
}

AMBIGUOUS_TERMS = {
    "station",
    "pageant",
    "authority",
    "society",
    "chamber",
    "mission",
    "forces",
    "business council",
    "international monetary fund",
    "career clinics",
    "prison",
    "settlement",
}


def is_confirmable(name: str, position: str, notes: str, source_excerpt: str) -> bool:
    if rejection_reason_for_name(name):
        return False
    lowered_blob = f" {position} {notes} {source_excerpt} ".lower()
    if any(term in lowered_blob for term in AMBIGUOUS_TERMS):
        return False
    if not any(term in lowered_blob for term in CONFIRMING_ROLE_TERMS):
        return False
    if "confidence: 95" not in lowered_blob and "confidence: 100" not in lowered_blob:
        return False
    return True


def main() -> None:
    backup_path = WORKBOOK_PATH.with_name(f"{WORKBOOK_PATH.stem}_pre_confirm_backup_{datetime.now():%Y%m%d_%H%M%S}{WORKBOOK_PATH.suffix}")
    shutil.copy2(WORKBOOK_PATH, backup_path)

    workbook = load_workbook(WORKBOOK_PATH)
    sheet = workbook["PEP Records"] if "PEP Records" in workbook.sheetnames else workbook.active
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    columns = {header: index + 1 for index, header in enumerate(headers)}

    confirmed = 0
    kept_review = 0
    today = datetime.now().strftime("%Y-%m-%d")
    for row in range(2, sheet.max_row + 1):
        status = str(sheet.cell(row, columns["status"]).value or "").strip()
        if status not in {"Needs review", "Candidate review"}:
            continue
        name = str(sheet.cell(row, columns["full_name"]).value or "").strip()
        position = str(sheet.cell(row, columns["position"]).value or "")
        notes = str(sheet.cell(row, columns["notes"]).value or "")
        source_excerpt = str(sheet.cell(row, columns["source_excerpt"]).value or "")
        source_name = str(sheet.cell(row, columns["source_name"]).value or "")
        if source_name != "Database.pdf":
            kept_review += 1
            continue
        if is_confirmable(name, position, notes, source_excerpt):
            sheet.cell(row, columns["status"]).value = "Confirmed"
            sheet.cell(row, columns["adverse_media_status"]).value = sheet.cell(row, columns["adverse_media_status"]).value or "No adverse media"
            sheet.cell(row, columns["last_reviewed_date"]).value = today
            reviewer_notes = str(sheet.cell(row, columns["reviewer_notes"]).value or "")
            add_note = "Confirmed from Database.pdf workbook review: direct public-role evidence and high confidence."
            sheet.cell(row, columns["reviewer_notes"]).value = f"{reviewer_notes}\n{today}: {add_note}".strip()
            confirmed += 1
        else:
            sheet.cell(row, columns["status"]).value = "Needs review"
            kept_review += 1

    workbook.save(WORKBOOK_PATH)
    print(f"Backup: {backup_path}")
    print(f"Rows confirmed: {confirmed}")
    print(f"Rows kept for review: {kept_review}")
    print(f"Updated workbook: {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
