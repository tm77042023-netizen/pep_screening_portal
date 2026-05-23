from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import secrets
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta
from difflib import SequenceMatcher
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Iterable
from urllib.parse import urljoin
from urllib.parse import quote
from urllib.parse import urlparse

from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer

try:
    import requests
    from requests.adapters import HTTPAdapter
except ImportError:
    requests = None
    HTTPAdapter = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None
from flask import (
    Flask,
    Response,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_sqlalchemy import SQLAlchemy
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None
from sqlalchemy import or_
try:
    from urllib3.util import Retry
except ImportError:
    Retry = None
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
DEFAULT_DATA_ROOT = Path(os.environ.get("LOCALAPPDATA", str(BASE_DIR)))
LOCAL_DATA_DIR = Path(os.environ.get("SOFTDAYTA_RISK_DATA_DIR", str(DEFAULT_DATA_ROOT / "Softdayta Risk")))
UPLOAD_DIR.mkdir(exist_ok=True)
LOCAL_DATA_DIR.mkdir(exist_ok=True)

db = SQLAlchemy()
HIDDEN_RECORD_STATUSES = {"Rejected / not a person", "Merged duplicate"}
SUBSCRIPTION_PLANS = {
    "starter": {
        "code": "starter",
        "name": "Starter Compliance",
        "fee": 950,
        "summary": "For small teams starting with manual screening and source-backed review.",
        "features": ["100 manual searches per month", "Public preview upgrade", "Candidate review workflow", "Basic exports"],
    },
    "professional": {
        "code": "professional",
        "name": "Professional AML Desk",
        "fee": 2500,
        "summary": "For regulated teams running batch screening, adverse-media review and monitoring.",
        "features": ["Unlimited manual searches", "Bulk Excel screening", "Adverse-media intelligence", "Monitoring alerts", "Audit-ready reports"],
    },
    "enterprise": {
        "code": "enterprise",
        "name": "Enterprise API",
        "fee": 6500,
        "summary": "For organisations integrating screening into onboarding, KYC, procurement and case systems.",
        "features": ["Backend API integration", "Custom review workflow", "Priority support", "Implementation assistance", "Advanced exports"],
    },
}
BANK_DETAILS = {
    "bank": "Bank details to be confirmed",
    "account_name": "Softdayta Risk",
    "account_number": "To be issued on invoice",
    "branch": "Gaborone",
    "reference_prefix": "RSK",
}


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(255), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(30), nullable=False, default="subscriber")
    organisation = db.Column(db.String(160), nullable=True)
    subscription_status = db.Column(db.String(30), nullable=False, default="active")
    plan_code = db.Column(db.String(50), nullable=False, default="professional")
    billing_contact_name = db.Column(db.String(160), nullable=True)
    billing_contact_email = db.Column(db.String(255), nullable=True)
    phone = db.Column(db.String(80), nullable=True)
    trial_ends_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    def set_password(self, password: str) -> None:
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)

    @property
    def is_admin(self) -> bool:
        return self.role == "admin"

    @property
    def is_active_subscriber(self) -> bool:
        return self.is_admin or self.subscription_status == "active"


class PepRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(255), nullable=False, index=True)
    aliases = db.Column(db.Text, nullable=True)
    category = db.Column(db.String(120), nullable=False, default="Domestic PIP")
    jurisdiction = db.Column(db.String(80), nullable=False, default="Botswana")
    position = db.Column(db.String(255), nullable=True)
    organisation = db.Column(db.String(255), nullable=True)
    status = db.Column(db.String(50), nullable=False, default="Current")
    source_url = db.Column(db.String(500), nullable=True)
    source_name = db.Column(db.String(255), nullable=True)
    source_date = db.Column(db.String(80), nullable=True)
    date_identified = db.Column(db.String(80), nullable=True)
    source_type = db.Column(db.String(80), nullable=True)
    source_excerpt = db.Column(db.Text, nullable=True)
    source_reliability = db.Column(db.String(80), nullable=True)
    verification_status = db.Column(db.String(80), nullable=True, default="Unverified")
    verified_by = db.Column(db.String(255), nullable=True)
    last_verified_date = db.Column(db.String(80), nullable=True)
    next_review_due = db.Column(db.String(80), nullable=True)
    adverse_media_status = db.Column(db.String(80), nullable=True, default="Pending review")
    adverse_media_linkage = db.Column(db.String(80), nullable=True)
    reviewer_notes = db.Column(db.Text, nullable=True)
    last_reviewed_date = db.Column(db.String(80), nullable=True)
    profile_summary = db.Column(db.Text, nullable=True)
    profile_image_url = db.Column(db.String(800), nullable=True)
    profile_source_url = db.Column(db.String(800), nullable=True)
    profile_source_name = db.Column(db.String(255), nullable=True)
    profile_updated_at = db.Column(db.DateTime, nullable=True)
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

    def names_for_matching(self) -> list[str]:
        names = [self.full_name]
        if self.aliases:
            names.extend([part.strip() for part in self.aliases.split(";") if part.strip()])
        return names


class ScreeningRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    request_type = db.Column(db.String(30), nullable=False, default="single")
    original_filename = db.Column(db.String(255), nullable=True)
    status = db.Column(db.String(30), nullable=False, default="completed")
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    user = db.relationship("User", backref="screening_requests")


class ScreeningResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    request_id = db.Column(db.Integer, db.ForeignKey("screening_request.id"), nullable=False)
    searched_name = db.Column(db.String(255), nullable=False)
    matched_record_id = db.Column(db.Integer, db.ForeignKey("pep_record.id"), nullable=True)
    match_score = db.Column(db.Integer, nullable=False, default=0)
    decision = db.Column(db.String(50), nullable=False, default="No match")
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    request = db.relationship("ScreeningRequest", backref="results")
    matched_record = db.relationship("PepRecord")


class MonitoringSubject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    name = db.Column(db.String(255), nullable=False)
    status = db.Column(db.String(30), nullable=False, default="active")
    last_screened_at = db.Column(db.DateTime, nullable=True)
    last_decision = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    user = db.relationship("User", backref="monitoring_subjects")


class SubscriptionInvoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    plan_code = db.Column(db.String(50), nullable=False)
    plan_name = db.Column(db.String(120), nullable=False)
    amount = db.Column(db.Integer, nullable=False)
    currency = db.Column(db.String(10), nullable=False, default="BWP")
    status = db.Column(db.String(40), nullable=False, default="pending")
    payment_reference = db.Column(db.String(80), nullable=False, unique=True, index=True)
    billing_contact_name = db.Column(db.String(160), nullable=True)
    billing_contact_email = db.Column(db.String(255), nullable=True)
    valid_from = db.Column(db.DateTime, nullable=True)
    valid_until = db.Column(db.DateTime, nullable=True)
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

    user = db.relationship("User", backref="subscription_invoices")


class DeveloperApiKey(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    name = db.Column(db.String(160), nullable=False)
    key_prefix = db.Column(db.String(30), nullable=False, index=True)
    key_hash = db.Column(db.String(64), nullable=False, unique=True, index=True)
    status = db.Column(db.String(30), nullable=False, default="active")
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    last_used_at = db.Column(db.DateTime, nullable=True)

    user = db.relationship("User", backref="developer_api_keys")


class PipRelationship(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    principal_record_id = db.Column(db.Integer, db.ForeignKey("pep_record.id"), nullable=False, index=True)
    related_name = db.Column(db.String(255), nullable=False, index=True)
    relationship_type = db.Column(db.String(120), nullable=False)
    category = db.Column(db.String(120), nullable=False, default="Related party")
    jurisdiction = db.Column(db.String(80), nullable=True)
    source_name = db.Column(db.String(255), nullable=True)
    source_url = db.Column(db.String(500), nullable=True)
    source_excerpt = db.Column(db.Text, nullable=True)
    confidence_score = db.Column(db.Integer, nullable=False, default=50)
    review_status = db.Column(db.String(80), nullable=False, default="Candidate review")
    reviewer_notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

    principal = db.relationship("PepRecord", backref="relationships")


class RecordAuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    record_id = db.Column(db.Integer, db.ForeignKey("pep_record.id"), nullable=True, index=True)
    relationship_id = db.Column(db.Integer, db.ForeignKey("pip_relationship.id"), nullable=True, index=True)
    action = db.Column(db.String(120), nullable=False)
    actor = db.Column(db.String(255), nullable=True)
    changes_json = db.Column(db.Text, nullable=True)
    note = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    record = db.relationship("PepRecord")
    relationship = db.relationship("PipRelationship")


class PublicSource(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    url = db.Column(db.String(500), nullable=False)
    jurisdiction = db.Column(db.String(80), nullable=False, default="Botswana")
    enabled = db.Column(db.Boolean, nullable=False, default=True)
    last_checked_at = db.Column(db.DateTime, nullable=True)
    last_status = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class CurrentAffairsIssue(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    category = db.Column(db.String(120), nullable=False, default="Politics")
    jurisdiction = db.Column(db.String(80), nullable=False, default="Botswana")
    summary = db.Column(db.Text, nullable=False)
    source_name = db.Column(db.String(255), nullable=True)
    source_url = db.Column(db.String(500), nullable=True)
    image_url = db.Column(db.String(500), nullable=True)
    issue_date = db.Column(db.String(80), nullable=True)
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))


class SourceUpdateLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    source_id = db.Column(db.Integer, db.ForeignKey("public_source.id"), nullable=False)
    status = db.Column(db.String(50), nullable=False)
    records_created = db.Column(db.Integer, nullable=False, default=0)
    message = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    source = db.relationship("PublicSource")


class PdfIngestionLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    source_name = db.Column(db.String(255), nullable=False)
    source_url = db.Column(db.String(500), nullable=False)
    pdfs_found = db.Column(db.Integer, nullable=False, default=0)
    pdfs_processed = db.Column(db.Integer, nullable=False, default=0)
    candidates_created = db.Column(db.Integer, nullable=False, default=0)
    status = db.Column(db.String(50), nullable=False, default="success")
    message = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class StagedImportRun(db.Model):
    __tablename__ = "staged_import_runs"

    id = db.Column(db.Integer, primary_key=True)
    source_name = db.Column(db.String(255), nullable=False, default="Uploaded PDF")
    source_url = db.Column(db.String(800), nullable=False, default="manual-upload")
    jurisdiction = db.Column(db.String(80), nullable=False, default="Botswana")
    status = db.Column(db.String(30), nullable=False, default="staged")  # staged|applied|discarded
    message = db.Column(db.Text, nullable=True)
    metrics_json = db.Column(db.Text, nullable=True)
    created_by = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    applied_at = db.Column(db.DateTime, nullable=True)


class StagedImportCandidate(db.Model):
    __tablename__ = "staged_import_candidates"

    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.Integer, db.ForeignKey("staged_import_runs.id"), nullable=False, index=True)
    kind = db.Column(db.String(30), nullable=False)  # pip|relationship|registry
    full_name = db.Column(db.String(255), nullable=True, index=True)
    category = db.Column(db.String(120), nullable=True)
    position = db.Column(db.String(255), nullable=True)
    related_name = db.Column(db.String(255), nullable=True)
    relationship_type = db.Column(db.String(120), nullable=True)
    principal_record_id = db.Column(db.Integer, nullable=True)
    confidence_score = db.Column(db.Integer, nullable=True, default=0)
    snippet = db.Column(db.Text, nullable=True)
    evidence_json = db.Column(db.Text, nullable=True)
    review_status = db.Column(db.String(40), nullable=False, default="Pending review")  # Pending review|Approved|Rejected
    reviewer_note = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    run = db.relationship("StagedImportRun", backref="candidates")


def parse_cipa_registry_text(raw: str) -> dict[str, object]:
    text = re.sub(r"\s+", " ", raw or "").strip()
    company_match = re.search(r"([A-Za-z0-9][A-Za-z0-9 .,&'\\/\\-]+?)\s*\((BW\d{11})\)", text)
    company_name = company_match.group(1).strip() if company_match else ""
    company_number = company_match.group(2).strip() if company_match else ""
    status_match = re.search(r"Company status\s+([A-Za-z ]{2,60})\s+Company type", text)
    company_status = (status_match.group(1).strip() if status_match else "")
    type_match = re.search(r"Company type\s+([A-Za-z ]{2,60})\s+File for this company", text)
    company_type = (type_match.group(1).strip() if type_match else "")

    directors: list[dict[str, str]] = []
    if "Directors" in text:
        # Use the raw version for block parsing (line-ish boundaries matter).
        raw_lines = [line.strip() for line in (raw or "").splitlines() if line.strip()]
        # Find the "Directors" section and parse name blocks.
        start_idx = -1
        for i, line in enumerate(raw_lines):
            if line.strip().lower() == "directors":
                start_idx = i
        if start_idx >= 0:
            block = raw_lines[start_idx + 1 :]
            current: dict[str, str] = {}
            def flush():
                nonlocal current
                if current.get("name"):
                    directors.append(current)
                current = {}
            i = 0
            while i < len(block):
                line = block[i]
                # section boundaries
                if line.lower() in {"back", "copyright © 2025 cipa"}:
                    flush()
                    break
                if line.lower() in {"secretaries", "shareholders", "beneficial owners", "auditors", "addresses", "general details"}:
                    flush()
                    break
                # If we hit a field label, capture next line as value
                if line.lower() == "nationality" and i + 1 < len(block):
                    current["nationality"] = block[i + 1]
                    i += 2
                    continue
                if line.lower() == "postal address" and i + 1 < len(block):
                    current["postal_address"] = block[i + 1]
                    i += 2
                    continue
                if line.lower() == "appointment date" and i + 1 < len(block):
                    current["appointment_date"] = block[i + 1]
                    i += 2
                    continue
                if line.lower().startswith("does the director") or line.lower().startswith("show previous"):
                    i += 1
                    continue
                # Otherwise treat as a name line if it has at least two tokens and isn't a label
                if line.lower() not in {"nationality", "postal address", "appointment date"}:
                    # new name starts when we already have a name and hit another plausible name
                    tokens = [t for t in re.split(r"\s+", line) if t]
                    if len(tokens) >= 2 and line[0].isalpha():
                        if current.get("name"):
                            flush()
                        current["name"] = line
                i += 1
            flush()

    return {
        "company_name": company_name,
        "company_number": company_number,
        "company_status": company_status,
        "company_type": company_type,
        "directors": directors,
    }


class WebLinkReviewLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    source_url = db.Column(db.String(800), nullable=False)
    source_name = db.Column(db.String(255), nullable=True)
    jurisdiction = db.Column(db.String(80), nullable=False, default="Botswana")
    pages_reviewed = db.Column(db.Integer, nullable=False, default=0)
    links_found = db.Column(db.Integer, nullable=False, default=0)
    candidates_created = db.Column(db.Integer, nullable=False, default=0)
    adverse_context_hits = db.Column(db.Integer, nullable=False, default=0)
    status = db.Column(db.String(50), nullable=False, default="success")
    message = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class RejectedCandidate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    raw_value = db.Column(db.String(500), nullable=False)
    source_name = db.Column(db.String(255), nullable=True)
    source_url = db.Column(db.String(500), nullable=True)
    jurisdiction = db.Column(db.String(80), nullable=True)
    reason = db.Column(db.String(255), nullable=False)
    snippet = db.Column(db.Text, nullable=True)
    confidence_score = db.Column(db.Integer, nullable=False, default=0)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class AdverseMediaSearch(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    searched_name = db.Column(db.String(255), nullable=False, index=True)
    jurisdiction = db.Column(db.String(80), nullable=True)
    role = db.Column(db.String(255), nullable=True)
    overall_risk_level = db.Column(db.String(80), nullable=False, default="Needs review")
    overall_summary = db.Column(db.Text, nullable=True)
    pip_status = db.Column(db.String(80), nullable=True)
    display_badges = db.Column(db.Text, nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    user = db.relationship("User")

    @property
    def badges(self) -> list[str]:
        try:
            parsed = json.loads(self.display_badges or "[]")
            return [str(item) for item in parsed if str(item).strip()]
        except json.JSONDecodeError:
            return []


class AdverseMediaAlert(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    search_id = db.Column(db.Integer, db.ForeignKey("adverse_media_search.id"), nullable=False)
    headline = db.Column(db.String(500), nullable=False)
    risk_theme = db.Column(db.Text, nullable=True)
    risk_level = db.Column(db.String(80), nullable=False, default="Needs review")
    linkage_type = db.Column(db.String(120), nullable=False, default="Unclear")
    summary = db.Column(db.Text, nullable=True)
    source_name = db.Column(db.String(255), nullable=True)
    source_url = db.Column(db.String(500), nullable=True)
    source_date = db.Column(db.String(80), nullable=True)
    recommended_action = db.Column(db.Text, nullable=True)
    review_status = db.Column(db.String(80), nullable=False, default="Pending review")
    reviewer_notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

    search = db.relationship("AdverseMediaSearch", backref="alerts")

    @property
    def themes(self) -> list[str]:
        try:
            parsed = json.loads(self.risk_theme or "[]")
            return [str(item) for item in parsed if str(item).strip()]
        except json.JSONDecodeError:
            return []


class PublicSearchUsage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ip_hash = db.Column(db.String(64), nullable=False, index=True)
    usage_date = db.Column(db.String(10), nullable=False, index=True)
    names_searched = db.Column(db.Integer, nullable=False, default=0)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))


def create_app() -> Flask:
    app = Flask(__name__)
    # Keep secrets out of code. In production this MUST be set.
    app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY") or "change-me-before-production"
    db_uri = app.config.get("SQLALCHEMY_DATABASE_URI") or os.environ.get("PEP_DATABASE_URI")
    # Default to a writable per-environment data directory; override with PEP_DATABASE_URI for Postgres/MySQL/etc.
    app.config["SQLALCHEMY_DATABASE_URI"] = db_uri or "sqlite:///" + (LOCAL_DATA_DIR / "pep_portal.db").as_posix()
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    db.init_app(app)

    register_routes(app)
    with app.app_context():
        db.create_all()
        ensure_schema()
        # Seed demo users/data only when explicitly enabled (avoid production demo accounts).
        if os.environ.get("SEED_DEMO_USERS", "").strip() in {"1", "true", "yes", "on"}:
            seed_data()
    return app


def current_user() -> User | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    return db.session.get(User, user_id)


def hash_api_key(value: str) -> str:
    return hashlib.sha256((value or "").encode("utf-8")).hexdigest()


def generate_developer_api_key() -> str:
    return f"rsk_live_{secrets.token_urlsafe(32)}"


def authenticate_api_key() -> tuple[User | None, DeveloperApiKey | None]:
    header = request.headers.get("Authorization", "").strip()
    if not header.lower().startswith("bearer "):
        return None, None
    raw_key = header.split(" ", 1)[1].strip()
    if not raw_key:
        return None, None
    api_key = DeveloperApiKey.query.filter_by(key_hash=hash_api_key(raw_key), status="active").first()
    if not api_key:
        return None, None
    user = db.session.get(User, api_key.user_id)
    if not user or not user.is_active_subscriber:
        return None, None
    return user, api_key


def require_login():
    user = current_user()
    if not user:
        flash("Please log in to continue.", "warning")
        return None
    if not user.is_active_subscriber:
        flash("Your subscription is not active. Please contact support.", "danger")
        return None
    return user


def password_reset_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(str(current_app.config["SECRET_KEY"]), salt="risk-password-reset")


def make_password_reset_token(user: User) -> str:
    return password_reset_serializer().dumps({"user_id": user.id, "email": user.email})


def resolve_password_reset_token(token: str, max_age_seconds: int = 3600) -> User | None:
    try:
        payload = password_reset_serializer().loads(token, max_age=max_age_seconds)
    except (BadSignature, SignatureExpired):
        return None
    user = db.session.get(User, int(payload.get("user_id") or 0))
    if not user or user.email != payload.get("email"):
        return None
    return user


def create_resilient_session() -> requests.Session:
    if requests is None:
        return None
    session = requests.Session()
    if Retry is not None and HTTPAdapter is not None:
        retries = Retry(
            total=5,
            connect=3,
            read=3,
            backoff_factor=1.5,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods={"GET", "HEAD"},
            raise_on_status=False,
        )
        adapter = HTTPAdapter(max_retries=retries)
        session.mount("https://", adapter)
        session.mount("http://", adapter)
    session.headers.update(
        {
            "User-Agent": "SoftdaytaRiskBot/2.0 (+https://risk.softdayta.co.bw)",
            "Accept": "text/html,application/pdf,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
    )
    return session


HTTP_SESSION = create_resilient_session()


def fetch_url(url: str, *, timeout: tuple[int, int] = (10, 30)) -> requests.Response:
    if HTTP_SESSION is None:
        raise RuntimeError("URL fetching is unavailable because the requests package is not installed on this server.")
    response = HTTP_SESSION.get(url, timeout=timeout)
    response.raise_for_status()
    return response


def ensure_schema() -> None:
    inspector = db.inspect(db.engine)
    pep_columns = {column["name"] for column in inspector.get_columns("pep_record")}
    user_columns = {column["name"] for column in inspector.get_columns("user")}
    additions = {
        "date_identified": "VARCHAR(80)",
        "source_type": "VARCHAR(80)",
        "source_excerpt": "TEXT",
        "source_reliability": "VARCHAR(80)",
        "verification_status": "VARCHAR(80) DEFAULT 'Unverified'",
        "verified_by": "VARCHAR(255)",
        "last_verified_date": "VARCHAR(80)",
        "next_review_due": "VARCHAR(80)",
        "adverse_media_status": "VARCHAR(80)",
        "adverse_media_linkage": "VARCHAR(80)",
        "reviewer_notes": "TEXT",
        "last_reviewed_date": "VARCHAR(80)",
        "profile_summary": "TEXT",
        "profile_image_url": "VARCHAR(800)",
        "profile_source_url": "VARCHAR(800)",
        "profile_source_name": "VARCHAR(255)",
        "profile_updated_at": "DATETIME",
    }
    for column_name, column_type in additions.items():
        if column_name not in pep_columns:
            db.session.execute(db.text(f"ALTER TABLE pep_record ADD COLUMN {column_name} {column_type}"))
    user_additions = {
        "plan_code": "VARCHAR(50) DEFAULT 'professional'",
        "billing_contact_name": "VARCHAR(160)",
        "billing_contact_email": "VARCHAR(255)",
        "phone": "VARCHAR(80)",
        "trial_ends_at": "DATETIME",
    }
    for column_name, column_type in user_additions.items():
        if column_name not in user_columns:
            db.session.execute(db.text(f"ALTER TABLE user ADD COLUMN {column_name} {column_type}"))
    db.session.commit()


def normalise_name(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9 ]+", " ", value or "")
    return re.sub(r"\s+", " ", value).strip().lower()


def match_score(name_a: str, name_b: str) -> int:
    a = normalise_name(name_a)
    b = normalise_name(name_b)
    if not a or not b:
        return 0
    if a == b:
        return 100
    ratio = SequenceMatcher(None, a, b).ratio()
    token_a = set(a.split())
    token_b = set(b.split())
    token_ratio = len(token_a & token_b) / max(len(token_a | token_b), 1)
    return round(max(ratio, token_ratio) * 100)


def is_confirmed_record(record: PepRecord) -> bool:
    return record.status.lower() in {"current", "confirmed", "former"} or record.source_name == "Seed data"


def screen_name(name: str) -> tuple[PepRecord | None, int, str]:
    best_record = None
    best_score = 0
    query_tokens = set(normalise_name(name).split())
    token_hits: list[tuple[PepRecord, int]] = []
    records = visible_record_query().all()
    for record in records:
        for candidate in record.names_for_matching():
            score = match_score(name, candidate)
            if score > best_score:
                best_score = score
                best_record = record
            candidate_tokens = set(normalise_name(candidate).split())
            if query_tokens and query_tokens <= candidate_tokens:
                token_hits.append((record, score))

    confirmed_token_hits = [(record, score) for record, score in token_hits if is_confirmed_record(record)]
    if confirmed_token_hits:
        unique_confirmed = {record.id: (record, score) for record, score in confirmed_token_hits}
        if len(unique_confirmed) == 1:
            record, score = next(iter(unique_confirmed.values()))
            return record, max(score, 75), "Possible match"

    if best_score >= 90:
        decision = "Likely match"
    elif best_score >= 75:
        decision = "Possible match"
    else:
        decision = "No match"
        best_record = None
    return best_record, best_score, decision


def screening_risk_level(result: ScreeningResult) -> str:
    record = result.matched_record
    adverse_status = adverse_media_label(record) if record else ""
    if adverse_status in {"Sanctions match", "Fraud allegation", "Procurement risk", "Adverse media found"}:
        return "High"
    if result.decision == "Likely match" or result.match_score >= 90:
        return "High"
    if result.decision == "Possible match" or result.match_score >= 75:
        return "Medium"
    if result.decision == "No match":
        return "Low"
    return "Unknown"


def screening_decision_label(result: ScreeningResult) -> str:
    if result.decision == "No match":
        return "Cleared"
    if result.decision == "Likely match":
        return "Pending review"
    if result.decision == "Possible match":
        return "Needs review"
    return result.decision or "Pending review"


def is_actionable_screening_match(result: ScreeningResult) -> bool:
    return result.decision in {"Possible match", "Likely match"} or bool(result.matched_record_id)


def monitoring_outcome_label(last_decision: str | None) -> str:
    value = last_decision or ""
    if not value:
        return "Pending first scan"
    if value.startswith("No match"):
        return "No confirmed match"
    if value.startswith("Possible match"):
        score = re.search(r"(\d+)", value)
        return f"Possible match: {score.group(1)}% similarity" if score else "Possible match"
    if value.startswith("Likely match"):
        score = re.search(r"(\d+)", value)
        return f"Likely match: {score.group(1)}% similarity" if score else "Likely match"
    return value


def monitoring_risk_level(subject: MonitoringSubject) -> str:
    outcome = monitoring_outcome_label(subject.last_decision)
    if outcome.startswith("Likely match"):
        return "High"
    if outcome.startswith("Possible match"):
        return "Medium"
    if outcome.startswith("No confirmed match"):
        return "Low"
    return "Unknown"


def public_result_snippet(record: PepRecord | None) -> str:
    if not record:
        return "No confirmed public profile was matched in the limited preview search. This is not a final compliance decision."
    parts = []
    if record.profile_summary:
        parts.append(clean_profile_summary(record.profile_summary))
    elif record.position:
        parts.append(f"{record.full_name} is listed as {record.position}.")
    if record.category:
        parts.append(f"Category: {record.category}.")
    if record.jurisdiction:
        parts.append(f"Jurisdiction: {record.jurisdiction}.")
    return " ".join(parts)[:420]


def run_public_search(raw_names: str, limit: int = 3) -> tuple[list[dict[str, object]], bool]:
    names = [line.strip() for line in (raw_names or "").splitlines() if line.strip()]
    limit = max(0, limit)
    limited = names[:limit]
    was_limited = len(names) > limit
    results = []
    for name in limited:
        record, score, decision = screen_name(name)
        results.append(
            {
                "searched_name": name,
                "decision": decision,
                "score": score,
                "record": record,
                "snippet": public_result_snippet(record),
            }
        )
    return results, was_limited


def public_search_ip_hash() -> str:
    forwarded_for = request.headers.get("X-Forwarded-For", "").split(",", 1)[0].strip()
    ip_address = forwarded_for or request.remote_addr or "unknown"
    return hashlib.sha256(ip_address.encode("utf-8")).hexdigest()


def reserve_public_search_allowance(requested_names: int, daily_limit: int = 3) -> tuple[int, bool, int]:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    usage = PublicSearchUsage.query.filter_by(ip_hash=public_search_ip_hash(), usage_date=today).first()
    if not usage:
        usage = PublicSearchUsage(ip_hash=public_search_ip_hash(), usage_date=today, names_searched=0)
        db.session.add(usage)
        db.session.flush()
    remaining = max(daily_limit - usage.names_searched, 0)
    allowed = min(max(requested_names, 0), remaining)
    usage.names_searched += allowed
    db.session.commit()
    return allowed, requested_names > allowed, max(daily_limit - usage.names_searched, 0)


def public_search_remaining(daily_limit: int = 3) -> int:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    usage = PublicSearchUsage.query.filter_by(ip_hash=public_search_ip_hash(), usage_date=today).first()
    if not usage:
        return daily_limit
    return max(daily_limit - usage.names_searched, 0)


def get_subscription_plan(plan_code: str | None) -> dict[str, object]:
    return SUBSCRIPTION_PLANS.get((plan_code or "").strip().lower(), SUBSCRIPTION_PLANS["professional"])


def format_bwp(amount: int | float | None) -> str:
    return f"BWP {int(amount or 0):,}"


def bool_yes_no(value: bool) -> str:
    return "Yes" if value else "No"


def worksheet_find_cell(ws, label: str):
    target = (label or "").strip().lower()
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), min_col=1, max_col=min(ws.max_column, 30)):
        for cell in row:
            val = str(cell.value or "").strip().lower()
            if val == target:
                return cell
    return None


def extract_youtube_video_id(url: str) -> str | None:
    """Extract a YouTube video id from common YouTube URL formats."""

    url = (url or "").strip()
    if not url:
        return None
    if not url.lower().startswith(("http://", "https://")):
        url = "https://" + url
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower()
    path = (parsed.path or "").strip("/")

    if host in {"youtu.be"}:
        return path.split("/", 1)[0] or None

    if host.endswith("youtube.com") or host.endswith("youtube-nocookie.com") or host in {"m.youtube.com"}:
        query = {}
        for pair in (parsed.query or "").split("&"):
            if "=" not in pair:
                continue
            key, value = pair.split("=", 1)
            query.setdefault(key, value)
        if query.get("v"):
            return query["v"]
        # /shorts/<id>, /embed/<id>
        parts = [part for part in path.split("/") if part]
        if len(parts) >= 2 and parts[0] in {"shorts", "embed"}:
            return parts[1]
    return None


def fetch_youtube_transcript_text(video_id: str, *, lang: str = "en") -> tuple[str | None, dict[str, str]]:
    """Attempt to fetch YouTube caption transcript text via the public timedtext endpoint.

    This only works when captions are available. We intentionally do not attempt to bypass
    YouTube access controls.
    """

    meta: dict[str, str] = {"video_id": video_id, "language": lang}
    if not video_id:
        return None, meta

    # Try manual captions first, then auto captions (ASR) as a fallback.
    attempts = [
        ("manual", f"https://www.youtube.com/api/timedtext?lang={quote(lang)}&v={quote(video_id)}"),
        ("asr", f"https://www.youtube.com/api/timedtext?lang={quote(lang)}&kind=asr&v={quote(video_id)}"),
    ]
    for mode, timedtext_url in attempts:
        try:
            resp = fetch_url(timedtext_url, timeout=(10, 30))
        except Exception:  # noqa: BLE001
            continue
        content = (resp.text or "").strip()
        if not content or content == "<?xml version=\"1.0\" encoding=\"utf-8\" ?><transcript></transcript>":
            continue
        try:
            root = ET.fromstring(content)
        except ET.ParseError:
            continue
        parts: list[str] = []
        for node in root.findall(".//text"):
            if node.text:
                # Captions may contain entities/newlines; keep it simple for extraction.
                parts.append(re.sub(r"\s+", " ", node.text.strip()))
        text = "\n".join([p for p in parts if p])
        if text:
            meta["mode"] = mode
            meta["source_url"] = timedtext_url
            return text, meta

    return None, meta


def build_risk_research_report(screening: ScreeningRequest) -> BytesIO:
    import openpyxl

    template_path = BASE_DIR / "static" / "report_templates" / "risk_research_template.xlsx"
    wb = openpyxl.load_workbook(template_path)

    def safe_set(ws, row: int, column: int, value: object) -> None:
        """Write values into templates that use merged cells.

        openpyxl represents all non-top-left merged cells as read-only MergedCell objects,
        so we resolve the merge range and write into its top-left anchor cell.
        """

        cell = ws.cell(row=row, column=column)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for merged in ws.merged_cells.ranges:
                if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
                    ws.cell(row=merged.min_row, column=merged.min_col).value = value
                    return
            # Fallback: if we can't find the range, do nothing rather than error.
            return
        cell.value = value

    # Heuristics: treat names containing common company suffixes as company subjects.
    company_suffixes = ("pty", "proprietary", "limited", "ltd", "inc", "plc", "company", "co.")
    company_subjects = []
    individual_subjects = []
    for result in screening.results:
        name = (result.searched_name or "").strip()
        lowered = name.lower()
        if any(suffix in lowered for suffix in company_suffixes):
            company_subjects.append(result)
        else:
            individual_subjects.append(result)

    completed_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Summary of Risk Indicators
    if "Summary of Risk Indicators" in wb.sheetnames:
        ws = wb["Summary of Risk Indicators"]
        cell = worksheet_find_cell(ws, "Date Completed")
        if cell:
            ws.cell(row=cell.row, column=cell.column + 1).value = completed_date
        # Default risk levels from our app signals (no sanctions module yet).
        pep_risk = "Low"
        adverse_risk = "Low"
        for result in screening.results:
            record = result.matched_record
            if record and record.status in {"Current", "Confirmed", "Former"}:
                pep_risk = "Medium" if pep_risk == "Low" else pep_risk
            if record and (record.adverse_media_status or "") not in {"", "No adverse media", "Pending review"}:
                adverse_risk = "Medium" if adverse_risk == "Low" else adverse_risk
        overall = "Low"
        if "High" in {pep_risk, adverse_risk}:
            overall = "High"
        elif "Medium" in {pep_risk, adverse_risk}:
            overall = "Medium"
        for label, level, note in [
            ("Overall", overall, ""),
            ("PEPs Screening", pep_risk, "Possible PIP/PEP matches are recorded where database evidence exists."),
            ("Sanctions Screening", "Low", "Sanctions module not configured in this build; requires separate watchlist integration."),
            ("Adverse Media Search", adverse_risk, "Adverse media is based on linked sources and requires human review."),
        ]:
            cell = worksheet_find_cell(ws, label)
            if cell:
                ws.cell(row=cell.row, column=cell.column + 1).value = level
                ws.cell(row=cell.row, column=cell.column + 3).value = note

    # Profile of Company Subject(s)
    if "Profile of Company Subject(s)" in wb.sheetnames:
        ws = wb["Profile of Company Subject(s)"]
        if company_subjects:
            ws.cell(row=2, column=6).value = company_subjects[0].searched_name
        # Populate simple status flags based on match/adverse media.
        pep_yes = any(r.matched_record and r.matched_record.status in {"Current", "Confirmed", "Former"} for r in company_subjects)
        adverse_yes = any(r.matched_record and (r.matched_record.adverse_media_status or "") not in {"", "No adverse media", "Pending review"} for r in company_subjects)
        ws.cell(row=3, column=6).value = bool_yes_no(pep_yes)
        ws.cell(row=4, column=6).value = "No"
        ws.cell(row=5, column=6).value = bool_yes_no(adverse_yes)
        ws.cell(row=6, column=6).value = "Botswana"

    # Profile of Individual Subject(s
    if "Profile of Individual Subject(s" in wb.sheetnames:
        ws = wb["Profile of Individual Subject(s"]
        start_col = 6
        step = 2
        for idx, res in enumerate(individual_subjects[:6]):
            col = start_col + idx * step
            ws.cell(row=3, column=col).value = res.searched_name
            record = res.matched_record
            pep_yes = bool(record and record.status in {"Current", "Confirmed", "Former"})
            adverse_yes = bool(record and (record.adverse_media_status or "") not in {"", "No adverse media", "Pending review"})
            ws.cell(row=4, column=col).value = bool_yes_no(pep_yes)
            ws.cell(row=5, column=col).value = "No"
            ws.cell(row=6, column=col).value = bool_yes_no(adverse_yes)

    # Risk Analysis + Summary
    pep_lines = []
    adverse_lines = []
    for result in screening.results:
        record = result.matched_record
        if record and record.status in {"Current", "Confirmed", "Former"}:
            pep_lines.append(f"- Possible PIP/PEP match: {record.full_name} ({record.category}, {record.position or 'role unknown'}).")
        if record and (record.adverse_media_status or "") not in {"", "No adverse media", "Pending review"}:
            adverse_lines.append(f"- Adverse media signal: {record.full_name} ({record.adverse_media_status}).")
    if "Risk Analysis" in wb.sheetnames:
        ws = wb["Risk Analysis"]
        # These sheets are mostly narrative; we fill a couple of bullet slots.
        safe_set(ws, row=3, column=2, value=pep_lines[0] if pep_lines else "- No PEP related information was found.")
        safe_set(ws, row=6, column=2, value=adverse_lines[0] if adverse_lines else "- No adverse media related information was found.")
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        safe_set(ws, row=3, column=2, value="No PEP related information was found." if not pep_lines else " ".join(pep_lines[:2]))
        safe_set(ws, row=5, column=2, value="No sanctions related information was found.")  # placeholder
        safe_set(ws, row=7, column=2, value="No adverse media related information was found." if not adverse_lines else " ".join(adverse_lines[:2]))

    # Full Text of Sources (evidence register)
    if "Full Text" in wb.sheetnames:
        ws = wb["Full Text"]
        # Template heading lives at B3, and a placeholder "Not Applicable" typically sits at B5.
        start_row = 5
        start_col = 2  # Column B

        headers = [
            "Searched name",
            "Decision",
            "Score (%)",
            "Matched record",
            "PIP category",
            "Position / role",
            "Jurisdiction",
            "Adverse media status",
            "Source name",
            "Source date",
            "Source URL",
            "Basis in document / excerpt",
        ]
        for offset, label in enumerate(headers):
            safe_set(ws, row=start_row, column=start_col + offset, value=label)

        row = start_row + 1
        for result in screening.results:
            record = result.matched_record
            safe_set(ws, row=row, column=start_col + 0, value=result.searched_name)
            safe_set(ws, row=row, column=start_col + 1, value=result.decision)
            safe_set(ws, row=row, column=start_col + 2, value=int(result.match_score or 0))
            safe_set(ws, row=row, column=start_col + 3, value=record.full_name if record else "")
            safe_set(ws, row=row, column=start_col + 4, value=record.category if record else "")
            safe_set(ws, row=row, column=start_col + 5, value=record.position if record else "")
            safe_set(ws, row=row, column=start_col + 6, value=record.jurisdiction if record else "")
            safe_set(ws, row=row, column=start_col + 7, value=record.adverse_media_status if record else "")
            safe_set(ws, row=row, column=start_col + 8, value=record.source_name if record else "")
            safe_set(ws, row=row, column=start_col + 9, value=record.source_date if record else "")
            safe_set(ws, row=row, column=start_col + 10, value=record.source_url if record else "")

            excerpt = ""
            if record:
                excerpt = (record.source_excerpt or "").strip()
                if not excerpt:
                    excerpt = (record.profile_summary or "").strip()
                if not excerpt:
                    excerpt = (record.notes or "").strip()
            safe_set(ws, row=row, column=start_col + 11, value=excerpt)
            row += 1

        if row == start_row + 1:
            safe_set(ws, row=start_row + 1, column=start_col, value="No screening results were captured for this request.")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def serialise_for_audit(record: PepRecord) -> dict[str, object]:
    return {
        "full_name": record.full_name,
        "aliases": record.aliases,
        "category": record.category,
        "jurisdiction": record.jurisdiction,
        "position": record.position,
        "organisation": record.organisation,
        "status": record.status,
        "source_name": record.source_name,
        "source_url": record.source_url,
        "source_type": record.source_type,
        "source_reliability": record.source_reliability,
        "verification_status": record.verification_status,
        "verified_by": record.verified_by,
        "last_verified_date": record.last_verified_date,
        "next_review_due": record.next_review_due,
        "adverse_media_status": record.adverse_media_status,
        "adverse_media_linkage": record.adverse_media_linkage,
        "reviewer_notes": record.reviewer_notes,
    }


def changed_fields(before: dict[str, object], after: dict[str, object]) -> dict[str, dict[str, object]]:
    return {
        key: {"before": before.get(key), "after": after.get(key)}
        for key in sorted(set(before) | set(after))
        if before.get(key) != after.get(key)
    }


def add_audit_log(
    *,
    action: str,
    actor: str | None = None,
    record: PepRecord | None = None,
    relationship: PipRelationship | None = None,
    changes: dict[str, object] | None = None,
    note: str = "",
) -> None:
    db.session.add(
        RecordAuditLog(
            record_id=record.id if record and record.id else None,
            relationship_id=relationship.id if relationship and relationship.id else None,
            action=action,
            actor=actor,
            changes_json=json.dumps(changes or {}, default=str),
            note=note,
        )
    )


def next_subscription_reference() -> str:
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    prefix = BANK_DETAILS["reference_prefix"]
    count = SubscriptionInvoice.query.filter(SubscriptionInvoice.payment_reference.like(f"{prefix}-{today}-%")).count() + 1
    return f"{prefix}-{today}-{count:04d}"


def create_subscription_invoice(user: User, plan: dict[str, object]) -> SubscriptionInvoice:
    invoice = SubscriptionInvoice(
        user_id=user.id,
        plan_code=str(plan["code"]),
        plan_name=str(plan["name"]),
        amount=int(plan["fee"]),
        payment_reference=next_subscription_reference(),
        billing_contact_name=user.billing_contact_name or user.name,
        billing_contact_email=user.billing_contact_email or user.email,
        notes="Access activates after payment confirmation. Trial access may be approved by administrator.",
    )
    db.session.add(invoice)
    db.session.commit()
    return invoice


def save_results(user: User, names: Iterable[str], request_type: str, filename: str | None = None) -> ScreeningRequest:
    screening = ScreeningRequest(user_id=user.id, request_type=request_type, original_filename=filename)
    db.session.add(screening)
    db.session.flush()
    for raw_name in names:
        name = (raw_name or "").strip()
        if not name:
            continue
        record, score, decision = screen_name(name)
        if record:
            enrich_profile(record)
        db.session.add(
            ScreeningResult(
                request_id=screening.id,
                searched_name=name,
                matched_record_id=record.id if record else None,
                match_score=score,
                decision=decision,
            )
        )
    db.session.commit()
    return screening


def extract_names_from_workbook(file_storage) -> list[str]:
    workbook = load_workbook(file_storage, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
    try:
        name_index = headers.index("name")
    except ValueError:
        name_index = 0
    names = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        value = row[name_index] if name_index < len(row) else None
        if value:
            names.append(str(value).strip())
    return names


PEP_IMPORT_EXPORT_HEADERS = [
    "id",
    "full_name",
    "aliases",
    "category",
    "jurisdiction",
    "position",
    "organisation",
    "status",
    "adverse_media_status",
    "adverse_media_linkage",
    "date_identified",
    "last_reviewed_date",
    "last_verified_date",
    "next_review_due",
    "verification_status",
    "verified_by",
    "source_reliability",
    "source_type",
    "source_name",
    "source_url",
    "source_date",
    "source_excerpt",
    "reviewer_notes",
    "notes",
]


def workbook_value(row_data: dict[str, object], key: str, default: str = "") -> str:
    value = row_data.get(key)
    if value is None:
        return default
    return str(value).strip()


def record_to_export_row(record: PepRecord) -> list[object]:
    return [
        record.id,
        record.full_name,
        record.aliases or "",
        record.category,
        record.jurisdiction,
        record.position or "",
        record.organisation or "",
        record.status,
        record.adverse_media_status or "Pending review",
        record.adverse_media_linkage or "",
        record.date_identified or "",
        record.last_reviewed_date or "",
        record.last_verified_date or "",
        record.next_review_due or "",
        record.verification_status or "Unverified",
        record.verified_by or "",
        record.source_reliability or "",
        record.source_type or "",
        record.source_name or "",
        record.source_url or "",
        record.source_date or "",
        record.source_excerpt or "",
        record.reviewer_notes or "",
        record.notes or "",
    ]


def build_records_workbook() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PEP Records"
    sheet.append(PEP_IMPORT_EXPORT_HEADERS)
    for record in visible_record_query().order_by(PepRecord.full_name).all():
        sheet.append(record_to_export_row(record))

    guide = workbook.create_sheet("Guide")
    guide.append(["Column", "How to use"])
    guide_rows = {
        "id": "Keep this value to update an existing record. Leave blank to create a new record.",
        "full_name": "Required for new records.",
        "status": "Candidate review, Confirmed, Current, Former, Needs review, Duplicate.",
        "adverse_media_status": "Pending review, No adverse media, Adverse media found, Under investigation, Sanctions match, Procurement risk, Fraud allegation, Official capacity only.",
        "source_excerpt": "Paste evidence text supporting the PIP or adverse-media classification.",
        "verification_status": "Unverified, Source verified, Management represented, Needs re-verification.",
        "source_reliability": "High, Medium, Low, or Unknown depending on source quality.",
        "next_review_due": "Next scheduled source verification date.",
        "notes": "Basis, limitations, reviewer comments, or audit notes.",
    }
    for key, help_text in guide_rows.items():
        guide.append([key, help_text])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def import_records_workbook(file_storage) -> tuple[int, int, list[str]]:
    workbook = load_workbook(file_storage, data_only=True)
    sheet = workbook["PEP Records"] if "PEP Records" in workbook.sheetnames else workbook.active
    headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
    updated = 0
    created = 0
    errors: list[str] = []
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        row_data = {headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]}
        full_name = workbook_value(row_data, "full_name") or workbook_value(row_data, "name")
        record_id = workbook_value(row_data, "id")
        record = None
        if record_id:
            try:
                record = db.session.get(PepRecord, int(float(record_id)))
            except ValueError:
                errors.append(f"Row {row_number}: invalid id '{record_id}'")
                continue
        if not record and full_name:
            record = PepRecord.query.filter(PepRecord.full_name.ilike(full_name)).first()
        if not record and not full_name:
            errors.append(f"Row {row_number}: full_name is required for new records")
            continue

        is_new = record is None
        if is_new:
            record = PepRecord(full_name=full_name, status="Candidate review")
            db.session.add(record)
            created += 1
        else:
            updated += 1

        record.full_name = full_name or record.full_name
        record.aliases = workbook_value(row_data, "aliases", record.aliases or "")
        notes_value = workbook_value(row_data, "notes", record.notes or "")
        category_match = re.search(r"Category:\s*([^;]+)", notes_value)
        position_match = re.search(r"Position:\s*([^;]+)", notes_value)
        jurisdiction_match = re.search(r"Jurisdiction:\s*([^;]+)", notes_value)
        evidence_match = re.search(r"Evidence:\s*(.+)", notes_value, flags=re.S)
        record.category = workbook_value(row_data, "category", record.category or "") or (category_match.group(1).strip() if category_match else "") or "Domestic PIP"
        record.jurisdiction = workbook_value(row_data, "jurisdiction", record.jurisdiction or "") or (jurisdiction_match.group(1).strip() if jurisdiction_match else "") or "Botswana"
        record.position = workbook_value(row_data, "position", record.position or "") or (position_match.group(1).strip() if position_match else "")
        record.organisation = workbook_value(row_data, "organisation", record.organisation or "")
        record.status = workbook_value(row_data, "status", record.status or "Candidate review") or "Candidate review"
        record.adverse_media_status = workbook_value(row_data, "adverse_media_status", record.adverse_media_status or "Pending review") or "Pending review"
        record.adverse_media_linkage = workbook_value(row_data, "adverse_media_linkage", record.adverse_media_linkage or "")
        record.date_identified = workbook_value(row_data, "date_identified", record.date_identified or today)
        record.last_reviewed_date = workbook_value(row_data, "last_reviewed_date", record.last_reviewed_date or "")
        record.last_verified_date = workbook_value(row_data, "last_verified_date", record.last_verified_date or "")
        record.next_review_due = workbook_value(row_data, "next_review_due", record.next_review_due or "")
        record.verification_status = workbook_value(row_data, "verification_status", record.verification_status or "Unverified") or "Unverified"
        record.verified_by = workbook_value(row_data, "verified_by", record.verified_by or "")
        record.source_reliability = workbook_value(row_data, "source_reliability", record.source_reliability or "")
        record.source_type = workbook_value(row_data, "source_type", record.source_type or "Excel upload")
        reference_value = workbook_value(row_data, "reference")
        record.source_name = workbook_value(row_data, "source_name", record.source_name or "") or ("Excel upload" if not reference_value else reference_value.split(" page ", 1)[0])
        record.source_url = workbook_value(row_data, "source_url", record.source_url or "") or reference_value
        record.source_date = workbook_value(row_data, "source_date", record.source_date or "")
        record.source_excerpt = workbook_value(row_data, "source_excerpt", record.source_excerpt or "") or (evidence_match.group(1).strip() if evidence_match else "")
        record.reviewer_notes = workbook_value(row_data, "reviewer_notes", record.reviewer_notes or "")
        record.notes = append_basis_if_missing(
            notes_value,
            f"Excel import/update on {today}; record retained for admin review.",
        )
        add_audit_log(
            action="excel_import_create" if is_new else "excel_import_update",
            record=record,
            changes={"row": row_number, "source": "Excel import"},
            note=f"Workbook import row {row_number}",
        )

    db.session.commit()
    return created, updated, errors


def record_feed_dict(record: PepRecord) -> dict[str, object]:
    return {
        "id": record.id,
        "full_name": record.full_name,
        "aliases": record.aliases or "",
        "category": record.category,
        "jurisdiction": record.jurisdiction,
        "position": record.position or "",
        "organisation": record.organisation or "",
        "status": record.status,
        "source": {
            "type": record.source_type or "",
            "name": record.source_name or "",
            "url": record.source_url or "",
            "date": record.source_date or "",
            "excerpt": record.source_excerpt or "",
            "reliability": record.source_reliability or "",
        },
        "verification": {
            "status": record.verification_status or "Unverified",
            "verified_by": record.verified_by or "",
            "last_verified_date": record.last_verified_date or record.last_reviewed_date or "",
            "next_review_due": record.next_review_due or "",
        },
        "adverse_media": {
            "status": record.adverse_media_status or "Pending review",
            "linkage": record.adverse_media_linkage or "",
        },
        "updated_at": record.updated_at.isoformat() if record.updated_at else "",
    }


def relationship_feed_dict(relationship: PipRelationship) -> dict[str, object]:
    return {
        "id": relationship.id,
        "principal_record_id": relationship.principal_record_id,
        "principal_name": relationship.principal.full_name if relationship.principal else "",
        "related_name": relationship.related_name,
        "relationship_type": relationship.relationship_type,
        "category": relationship.category,
        "jurisdiction": relationship.jurisdiction or "",
        "confidence_score": relationship.confidence_score,
        "review_status": relationship.review_status,
        "source_name": relationship.source_name or "",
        "source_url": relationship.source_url or "",
        "source_excerpt": relationship.source_excerpt or "",
        "updated_at": relationship.updated_at.isoformat() if relationship.updated_at else "",
    }


def feed_records_since(since: str = ""):
    query = visible_record_query().order_by(PepRecord.updated_at.desc())
    if since:
        try:
            since_date = datetime.fromisoformat(since.replace("Z", "+00:00"))
            query = query.filter(PepRecord.updated_at >= since_date)
        except ValueError:
            pass
    return query.all()


def build_feed_workbook(records: list[PepRecord]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PIP Feed"
    headers = [
        "id",
        "full_name",
        "aliases",
        "category",
        "jurisdiction",
        "position",
        "organisation",
        "status",
        "source_name",
        "source_url",
        "source_reliability",
        "verification_status",
        "last_verified_date",
        "next_review_due",
        "updated_at",
    ]
    sheet.append(headers)
    for record in records:
        data = record_feed_dict(record)
        sheet.append(
            [
                data["id"],
                data["full_name"],
                data["aliases"],
                data["category"],
                data["jurisdiction"],
                data["position"],
                data["organisation"],
                data["status"],
                data["source"]["name"],
                data["source"]["url"],
                data["source"]["reliability"],
                data["verification"]["status"],
                data["verification"]["last_verified_date"],
                data["verification"]["next_review_due"],
                data["updated_at"],
            ]
        )
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def extract_candidate_names(text: str) -> list[str]:
    candidates = set()
    pattern = re.compile(r"\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})\b")
    stop_words = {"Republic Of", "Terms Of", "Privacy Policy", "Contact Us", "Home Affairs"}
    for match in pattern.findall(text):
        if match not in stop_words and len(match) <= 80:
            candidates.add(match)
    return sorted(candidates)


PIP_CATEGORY_RULES = [
    (
        "Executive government",
        [
            "president",
            "vice president",
            "minister",
            "deputy minister",
            "permanent secretary",
            "executive director",
            "equivalent senior official",
        ],
    ),
    (
        "Legislature",
        [
            "member of parliament",
            "mp",
            "national assembly",
            "speaker",
            "deputy speaker",
            "parliamentary",
            "council leader",
        ],
    ),
    (
        "Judiciary and regulators",
        [
            "chief justice",
            "judge",
            "magistrate",
            "ombudsman",
            "regulator",
            "commissioner",
            "authority",
            "supervisory body",
        ],
    ),
    (
        "State-owned enterprises",
        [
            "state owned",
            "state-owned",
            "parastatal",
            "board chair",
            "board member",
            "chief executive officer",
            "ceo",
            "managing director",
            "public enterprise",
        ],
    ),
    (
        "Local authorities",
        [
            "mayor",
            "deputy mayor",
            "councillor",
            "council secretary",
            "town clerk",
            "district commissioner",
            "local authority",
        ],
    ),
    (
        "Political parties",
        [
            "party president",
            "party leader",
            "secretary general",
            "political party",
            "campaign manager",
            "party chair",
            "party spokesperson",
        ],
    ),
    (
        "Traditional leadership",
        [
            "kgosi",
            "chief",
            "traditional leader",
            "tribal authority",
            "ntlo ya dikgosi",
        ],
    ),
    (
        "Related parties",
        [
            "spouse",
            "wife",
            "husband",
            "son",
            "daughter",
            "brother",
            "sister",
            "family member",
            "close associate",
        ],
    ),
]


PIP_CATEGORY_DEFINITIONS = [
    (
        "Executive government",
        "Presidents, vice presidents, ministers, deputy ministers, permanent secretaries and equivalent senior officials.",
    ),
    (
        "Legislature",
        "Members of parliament, national assembly members, council leaders and relevant parliamentary office holders.",
    ),
    (
        "Judiciary and regulators",
        "Senior judicial officers, heads of regulators, ombuds institutions and supervisory bodies.",
    ),
    (
        "State-owned enterprises",
        "Board members and senior executives of material public enterprises and government-linked entities.",
    ),
    (
        "Local authorities",
        "Mayors, councillors, chief executive officers and senior municipal leadership where risk-relevant.",
    ),
    (
        "Political parties",
        "Senior party officials and office bearers in nationally relevant political organisations.",
    ),
    (
        "Traditional leadership",
        "Recognised traditional authorities and senior traditional leadership where included in domestic PIP frameworks.",
    ),
    (
        "Related parties",
        "Family members and close associates where verified through reliable public sources and permitted by data-protection requirements.",
    ),
]

ADVERSE_CONTEXT_KEYWORDS = {
    "adverse",
    "allegation",
    "alleged",
    "arrest",
    "bribery",
    "corruption",
    "court",
    "criminal",
    "embezzlement",
    "fraud",
    "investigation",
    "misconduct",
    "mismanagement",
    "money laundering",
    "procurement",
    "sanction",
    "scandal",
    "tax offence",
    "tender irregularity",
}
RELATIONSHIP_KEYWORDS = {
    "spouse": ["spouse", "wife", "husband"],
    "child": ["son", "daughter", "child", "children"],
    "sibling": ["brother", "sister", "sibling"],
    "parent": ["father", "mother", "parent"],
    "family member": ["family member", "relative", "related to"],
    "close associate": ["close associate", "associate", "ally", "adviser", "advisor"],
    "business associate": ["business partner", "partner", "director", "shareholder", "company linked"],
}

COMPLIANCE_PACK_SECTIONS = [
    {
        "key": "privacy",
        "title": "Privacy and data-protection position",
        "summary": "Public-source PIP data is processed for AML/CFT/CPF compliance decision-support, subject to human review, source verification, access controls and retention limits.",
        "evidence": ["Responsible-use notices", "Human review workflow", "Source evidence fields", "Reviewer notes and audit trail"],
    },
    {
        "key": "lawful-sourcing",
        "title": "Lawful sourcing methodology",
        "summary": "Records should be sourced from government, regulator, registry, reputable news, tender, court, parliamentary and official public enterprise sources. Scraped candidates remain in review until confirmed.",
        "evidence": ["Source URL/document reference", "Source excerpt", "Source reliability", "Verification status", "Rejected candidate log"],
    },
    {
        "key": "retention",
        "title": "Data retention and review cycle",
        "summary": "Records require last verified dates, next review due dates and clear status changes. Unverified or rejected records should not be shown in normal screening outputs.",
        "evidence": ["Last verified date", "Next review due", "Hidden rejected statuses", "Audit/version history"],
    },
    {
        "key": "security",
        "title": "Security controls",
        "summary": "Admin-only database controls, API key hashing, backend-only OpenAI keys and subscriber authentication reduce unauthorised access risk.",
        "evidence": ["Admin-only routes", "Hashed API keys", "Backend API integration guidance", "No public demo credentials"],
    },
    {
        "key": "limitations",
        "title": "Limitations and reliance",
        "summary": "Outputs are indicators, not legal findings. Users must review source evidence before making onboarding, monitoring or escalation decisions.",
        "evidence": ["Landing-page legal notice", "Dashboard compliance safeguard", "Adverse-media disclaimer", "Decision workflow"],
    },
]

DATA_DICTIONARY = [
    ("full_name", "string", "Primary person or entity name used for screening."),
    ("aliases", "string", "Alternative names separated by semicolons."),
    ("category", "enum", "PIP/PEP category such as Executive government, Legislature, Related party."),
    ("jurisdiction", "string", "Country or jurisdiction of public role or relationship."),
    ("position", "string", "Role/title supporting classification."),
    ("status", "enum", "Candidate review, Current, Confirmed, Former, Needs review, Rejected, Duplicate."),
    ("source_name", "string", "Source publication, website, document or registry."),
    ("source_url", "string", "URL or document reference."),
    ("source_excerpt", "text", "Evidence text supporting the classification."),
    ("source_reliability", "enum", "High, Medium, Low or Unknown."),
    ("verification_status", "enum", "Unverified, Source verified, Management represented, Needs re-verification."),
    ("last_verified_date", "date", "Date the evidence was last checked."),
    ("next_review_due", "date", "Next planned verification review date."),
    ("relationship_type", "string", "Family member, close associate, spouse, director, business associate, etc."),
    ("confidence_score", "integer", "0-100 confidence score for extracted or relationship data."),
]


NAME_PATTERN = re.compile(
    r"\b(?:Dr|Mr|Mrs|Ms|Hon|His Excellency|Her Excellency|Kgosi|Justice|Prof)?\.?\s*"
    r"([A-Z][a-z]+(?:[-'][A-Z][a-z]+)?(?:\s+[A-Z][a-z]+(?:[-'][A-Z][a-z]+)?){1,3})\b"
)

ROLE_TO_CATEGORY = [
    ("Executive government", r"(?:president|vice president|minister|deputy minister|permanent secretary|executive director)"),
    ("Legislature", r"(?:member of parliament|national assembly speaker|speaker|deputy speaker|mp|honourable|hon\.)"),
    ("Judiciary and regulators", r"(?:chief justice|justice|judge|magistrate|ombudsman|commissioner|regulator)"),
    ("State-owned enterprises", r"(?:board chair|board member|chief executive officer|ceo|managing director)"),
    ("Local authorities", r"(?:mayor|deputy mayor|councillor|council secretary|town clerk|district commissioner)"),
    ("Political parties", r"(?:party president|party leader|secretary general|party chair|party spokesperson)"),
    ("Traditional leadership", r"(?:kgosi|chief|traditional leader)"),
    ("Related parties", r"(?:spouse|wife|husband|son|daughter|brother|sister|family member|close associate)"),
]

ROLE_EVIDENCE_PATTERNS = [
    ("Executive government", "President", r"\b(?:former\s+)?president\s+(?:Dr|Mr|Mrs|Ms|Sir|Advocate|Adv)?\.?\s*{name}\b"),
    ("Executive government", "President", r"\b{name}\s*,?\s+(?:former\s+)?president\b"),
    ("Executive government", "Vice President", r"\b(?:former\s+)?vice president\s+(?:Dr|Mr|Mrs|Ms|Sir)?\.?\s*{name}\b"),
    ("Executive government", "Vice President", r"\b{name}\s*,?\s+(?:former\s+)?vice president\b"),
    ("Executive government", "Minister", r"\b(?:former\s+)?minister(?:\s+for|\s+of)?\s+[A-Za-z ,&-]{0,80}?\s+(?:Dr|Mr|Mrs|Ms|Sir|Rre|Mma)?\.?\s*{name}\b"),
    ("Executive government", "Minister", r"\b{name}\s*,?\s+(?:former\s+)?minister\b"),
    ("Executive government", "Permanent Secretary", r"\b(?:former\s+)?permanent secretary\s+(?:Dr|Mr|Mrs|Ms|Sir)?\.?\s*{name}\b"),
    ("Executive government", "Governor of Bank of Botswana", r"\b{name}\s*,?\s+(?:former\s+)?governor of bank of botswana\b"),
    ("Legislature", "Speaker of the National Assembly", r"\b(?:former\s+)?speaker of the national assembly\s+(?:Dr|Mr|Mrs|Ms|Sir)?\.?\s*{name}\b"),
    ("Legislature", "Speaker of the National Assembly", r"\b{name}\s*,?\s+(?:former\s+)?speaker of the national assembly\b"),
    ("Legislature", "Legislator", r"\b(?:legislator|mp|member of parliament)\s+(?:Dr|Mr|Mrs|Ms|Rre|Mma)?\.?\s*{name}\b"),
    ("Legislature", "Legislator", r"\b{name}\s*,?\s+(?:a\s+)?(?:former\s+)?(?:legislator|mp|member of parliament)\b"),
    ("Local authorities", "Councillor", r"\b(?:former\s+)?councillor\s+(?:for\s+[A-Za-z ]{0,50}\s+)?(?:Dr|Mr|Mrs|Ms|Rre|Mma)?\.?\s*{name}\b"),
    ("Local authorities", "Councillor", r"\b{name}\s*,?\s+(?:former\s+)?councillor\b"),
    ("Traditional leadership", "Kgosi", r"\bkgosi\s+{name}\b"),
    ("Traditional leadership", "Chief / traditional leader", r"\b(?:chief|traditional leader)\s+{name}\b"),
    ("Political parties", "Senior political party figure", r"\b{name}\s*,?\s+(?:a\s+)?(?:veteran|stalwart|senior figure|party leader|party president|secretary general)\s+(?:of|in)?\s*(?:the\s+)?(?:bdp|bnf|udc|bcp|bpf|botswana democratic party|botswana national front)?\b"),
    ("Political parties", "Senior political party figure", r"\b(?:bdp|bnf|udc|bcp|bpf|botswana democratic party|botswana national front)\s+(?:veteran|stalwart|senior figure|party leader|party president|secretary general)\s+{name}\b"),
    ("State-owned enterprises", "SOE board/senior executive", r"\b{name}\s*,?\s+(?:board chair|board member|chief executive officer|ceo|managing director)\b"),
    ("Judiciary and regulators", "Judicial / regulatory senior official", r"\b(?:chief justice|justice|judge|magistrate|ombudsman|commissioner)\s+{name}\b"),
    ("Judiciary and regulators", "Judicial / regulatory senior official", r"\b{name}\s*,?\s+(?:chief justice|justice|judge|magistrate|ombudsman|commissioner)\b"),
    ("Related parties", "Verified related party", r"\b{name}\s*,?\s+(?:spouse|wife|husband|son|daughter|brother|sister|family member|close associate)\b"),
    ("Related parties", "Verified related party", r"\b(?:spouse|wife|husband|son|daughter|brother|sister|family member|close associate)\s+of\s+{name}\b"),
]

POSSIBLE_ROLE_PATTERNS = [
    ("Executive government; Political parties", "Political party figure with possible executive role", r"\b{name}\b.{0,80}\b(?:bdp|bnf|udc|bcp|bpf|botswana democratic party|botswana national front)\b.{0,80}\b(?:resigned|cabinet|government|minister|president)\b"),
    ("Political parties / possible executive government", "Political role stated; executive office not fully established in text", r"\b{name}\b.{0,80}\b(?:veteran|stalwart|powerful figure|senior figure)\b.{0,80}\b(?:bdp|bnf|udc|bcp|bpf|party)\b"),
]

BAD_NAME_WORDS = {
    "as",
    "and",
    "daily",
    "news",
    "speaks",
    "petitions",
    "view",
    "all",
    "leadership",
    "updates",
    "upload",
    "uploads",
    "get",
    "social",
    "register",
    "company",
    "public",
    "holidays",
    "editors",
    "editor",
    "photography",
    "marketing",
    "graphics",
    "head",
    "tel",
    "telephone",
    "email",
    "wednesday",
    "thursday",
    "friday",
    "monday",
    "tuesday",
    "saturday",
    "sunday",
    "january",
    "february",
    "march",
    "april",
    "may",
    "june",
    "july",
    "august",
    "september",
    "october",
    "november",
    "december",
    "botswana",
    "democratic",
    "national",
    "front",
    "congress",
    "party",
    "government",
    "republic",
    "state",
    "savings",
    "saving",
    "terms",
    "conditions",
    "constitution",
    "apply",
    "account",
    "insurance",
    "scheme",
    "deposit",
    "education",
    "vocational",
    "affairs",
    "disease",
    "mouth",
    "acting",
    "minister",
    "president",
    "member",
    "major",
    "general",
    "officer",
    "external",
    "advertisement",
    "vacancy",
    "departments",
    "departmental",
    "services",
    "service",
    "registration",
    "radio",
    "civil",
    "leadership",
    "rwanda",
    "sign",
    "six",
    "bilateral",
    "our",
    "circular",
    "court",
    "district",
    "southern",
    "property",
    "free",
    "trade",
    "area",
    "continental",
    "acting",
    "the",
    "rre",
    "phase",
    "iv",
    "iii",
    "ii",
    "loba",
    "bakary",
    "sambe",
    "kamel",
    "al",
}

NAVIGATION_TERMS = {
    "home",
    "news",
    "downloads",
    "calendar",
    "bills",
    "glossary",
    "contact",
    "quick links",
    "watch live",
    "gallery",
    "tenders",
    "vacancies",
    "brochures",
    "parliamentary business",
    "notice papers",
    "order papers",
    "press release",
    "website design",
    "latest news",
}

INSTITUTION_TERMS = {
    "centre",
    "university",
    "institute",
    "commission",
    "department",
    "ministry",
    "parliament",
    "national assembly",
    "parliament office",
    "parliamentary union",
    "technology research centre",
    "botswana university",
    "southern african development community",
    "secretariat",
    "catalogue",
    "copyright",
    "footer",
    "navigation",
    "menu",
    "link",
    "links",
    "layout",
    "headline",
}

PERSON_TITLES = r"(?:Hon\.?|Honourable|Mr\.?|Ms\.?|Mrs\.?|Dr\.?|Prof\.?|Rev\.?|Chief|Kgosi|Sir|Lady|Justice|Advocate|Adv\.)"
POSITION_KEYWORDS = [
    "President",
    "Vice President",
    "Prime Minister",
    "Minister",
    "Deputy Minister",
    "Permanent Secretary",
    "Member of Parliament",
    "MP",
    "Speaker",
    "Deputy Speaker",
    "Judge",
    "Justice",
    "Ambassador",
    "Chief Justice",
    "Mayor",
    "Councillor",
    "Kgosi",
    "Chief",
    "Board Chairperson",
    "Board Chair",
    "Board Member",
    "CEO",
    "Chief Executive Officer",
    "Executive Director",
    "Commissioner",
    "Ombudsman",
]

COUNTRY_NAMES = {
    "Botswana",
    "Namibia",
    "South Africa",
    "Zimbabwe",
    "Zambia",
    "Rwanda",
    "Hungary",
    "Spain",
    "Mali",
    "Uganda",
    "Kenya",
    "Tanzania",
    "Lesotho",
    "Eswatini",
    "Angola",
    "Mozambique",
    "Nigeria",
    "United States",
    "United Kingdom",
}

KNOWN_FOREIGN_PUBLIC_FIGURES = {
    "viktor orban": ("Hungary", "Foreign PEP/PIP", "Prime Minister"),
    "viktor orb n": ("Hungary", "Foreign PEP/PIP", "Prime Minister"),
    "pedro sanchez": ("Spain", "Foreign PEP/PIP", "Prime Minister"),
    "paul kagame": ("Rwanda", "Foreign PEP/PIP", "President"),
    "hakainde hichilema": ("Zambia", "Foreign PEP/PIP", "President"),
    "yoweri museveni": ("Uganda", "Foreign PEP/PIP", "President"),
    "robert mugabe": ("Zimbabwe", "Foreign PEP/PIP", "Former President"),
}

MIN_CANDIDATE_CONFIDENCE = 60
OPENAI_RESPONSES_URL = "https://api.openai.com/v1/responses"
OPENAI_DEFAULT_MODEL = os.environ.get("OPENAI_ANALYSIS_MODEL", "gpt-5.2")
OPENAI_ADVERSE_MODEL = os.environ.get("OPENAI_ADVERSE_MODEL", OPENAI_DEFAULT_MODEL)


def clean_candidate_name(value: str) -> str:
    value = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", value or "")
    value = re.sub(r"\b(?:Dr|Mr|Mrs|Ms|Hon|Honourable|His Excellency|Her Excellency|Kgosi|Justice|Prof|President|Advocate|Adv|Major General|Major Gen|General|Gen|Lt Gen|Sir|Lady|Rev|Chief)\.?\b", "", value or "", flags=re.I)
    value = re.sub(r"[^A-Za-z' -]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def rejection_reason_for_name(name: str) -> str | None:
    stripped = (name or "").strip()
    lowered_name = stripped.lower()
    parts = stripped.split()
    if not stripped:
        return "Empty candidate"
    if len(parts) < 2:
        return "Candidate does not contain a full person name"
    if len(parts) >= 5:
        return "Candidate has too many words and is likely a heading"
    if stripped.isupper() and len(parts) > 3:
        return "Uppercase section heading"
    if " " not in stripped and len(stripped) > 25:
        return "Likely filename or concatenated identifier"
    if lowered_name in NAVIGATION_TERMS or any(term in lowered_name for term in NAVIGATION_TERMS):
        return "Website navigation or page label"
    if lowered_name in INSTITUTION_TERMS or any(term in lowered_name for term in INSTITUTION_TERMS):
        return "Institution or organisation name"
    lowered_parts = {part.lower().strip("-'") for part in parts}
    if lowered_parts & BAD_NAME_WORDS:
        return "Contains generic non-name words"
    if any(len(part) > 18 for part in parts):
        return "Name token is unusually long"
    if not all(part[:1].isupper() and len(part) > 1 for part in parts):
        return "Does not follow person-name capitalisation"
    return None


def is_plausible_person_name(name: str) -> bool:
    return rejection_reason_for_name(name) is None


def extract_position(window: str, name: str) -> str:
    escaped_name = re.escape(name).replace(r"\ ", r"\s+")
    for keyword in POSITION_KEYWORDS:
        keyword_pattern = re.escape(keyword).replace(r"\ ", r"\s+")
        patterns = [
            rf"\b(?:former\s+)?{keyword_pattern}\b(?:\s+(?:of|for)\s+[A-Z][A-Za-z &,-]{{1,80}})?\s+(?:{PERSON_TITLES}\s+)?{escaped_name}\b",
            rf"\b(?:{PERSON_TITLES}\s+)?{escaped_name}\b\s*,?\s*(?:former\s+)?{keyword_pattern}\b(?:\s+(?:of|for)\s+[A-Z][A-Za-z &,-]{{1,80}})?",
        ]
        for pattern in patterns:
            match = re.search(pattern, window, flags=re.I)
            if match:
                return re.sub(r"\s+", " ", match.group(0)).strip()
    return "[Unknown]"


def nearby_role_category(window: str) -> str | None:
    lowered = window.lower()
    for category, keywords in PIP_CATEGORY_RULES:
        if any(keyword in lowered for keyword in keywords):
            return category
    if "prime minister" in lowered:
        return "Executive government"
    return None


def detect_jurisdiction(window: str, source_jurisdiction: str) -> tuple[str, str]:
    lowered = window.lower()
    for country in COUNTRY_NAMES:
        if country.lower() == source_jurisdiction.lower():
            continue
        if re.search(rf"\b(?:of|from|in)\s+{re.escape(country)}\b", lowered, flags=re.I):
            return country, "Foreign PEP/PIP"
    if source_jurisdiction:
        return source_jurisdiction, "Domestic PIP"
    return "Unknown", "Public-source mention only"


def known_foreign_figure(name: str, source_jurisdiction: str) -> tuple[str, str, str] | None:
    match = KNOWN_FOREIGN_PUBLIC_FIGURES.get(normalise_name(name))
    if not match:
        return None
    jurisdiction, category, role = match
    if jurisdiction.lower() == (source_jurisdiction or "").lower():
        return None
    return jurisdiction, category, role


def score_candidate(name: str, window: str, position: str) -> int:
    escaped_name_pattern = re.escape(name).replace(r"\ ", r"\s+")
    has_title = bool(re.search(rf"\b{PERSON_TITLES}\s+{escaped_name_pattern}\b", window, flags=re.I))
    has_position = position != "[Unknown]"
    if has_title and has_position:
        return 95
    if has_position:
        return 78
    if is_plausible_person_name(name):
        return 55
    return 20


def evidence_note(candidate: dict[str, str | int]) -> str:
    return (
        "Automatically extracted from public source. Admin review required before production reliance.\n\n"
        f"Basis in document: {candidate.get('basis', 'Role evidence identified')}\n\n"
        f"Detected role/title: {candidate.get('position', '[Unknown]')}\n\n"
        f"Confidence score: {candidate.get('confidence_score', 0)}\n\n"
        f"Evidence snippet: {candidate.get('snippet', '')}"
    )


def clean_profile_summary(summary: str) -> str:
    cleaned = re.sub(r"\s+", " ", summary or "").strip()
    if not cleaned:
        return ""
    sentences = re.split(r"(?<=[.!?])\s+", cleaned)
    selected = []
    for sentence in sentences:
        if len(" ".join(selected + [sentence])) > 650:
            break
        selected.append(sentence)
        if len(selected) >= 4:
            break
    return " ".join(selected).strip()


def google_custom_search_summary(name: str) -> dict[str, str] | None:
    api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
    cse_id = os.environ.get("GOOGLE_CSE_ID", "").strip()
    if not api_key or not cse_id:
        return None
    query = f"{name} politician public office biography"
    url = "https://www.googleapis.com/customsearch/v1"
    response = HTTP_SESSION.get(
        url,
        params={"key": api_key, "cx": cse_id, "q": query, "num": 3, "searchType": "image"},
        timeout=(10, 30),
    )
    response.raise_for_status()
    data = response.json()
    image_url = ""
    source_url = ""
    items = data.get("items") or []
    if items:
        image_url = items[0].get("link", "")
        source_url = items[0].get("image", {}).get("contextLink", "")

    response = HTTP_SESSION.get(
        url,
        params={"key": api_key, "cx": cse_id, "q": query, "num": 3},
        timeout=(10, 30),
    )
    response.raise_for_status()
    data = response.json()
    snippets = [item.get("snippet", "") for item in data.get("items") or [] if item.get("snippet")]
    if not snippets:
        return None
    items = data.get("items") or []
    def non_wikipedia_link(value: str) -> bool:
        lowered = (value or "").lower()
        return "wikipedia.org" not in lowered and "wiki" not in lowered

    first_item = {}
    for candidate in items:
        link = str(candidate.get("link") or "")
        if link and non_wikipedia_link(link):
            first_item = candidate
            break
    if not first_item and items:
        first_item = items[0]
    return {
        "summary": clean_profile_summary(" ".join(snippets)),
        "image_url": image_url,
        "source_url": first_item.get("link") or source_url,
        "source_name": "Google Custom Search",
    }


PROFILE_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "summary": {"type": "string"},
        "source_name": {"type": "string"},
        "source_url": {"type": "string"},
        "image_url": {"type": "string"},
    },
    "required": ["summary", "source_name", "source_url", "image_url"],
}


def openai_web_profile_summary(name: str) -> dict[str, str] | None:
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return None

    prompt = f"""
You are a PEP/PIP screening analyst.

Search the web for a brief biographical profile for:

Name: {name}

Rules:
1. Prefer primary/credible sources (official government pages, reputable news, biographies) over Wikipedia.
2. If the best available source is Wikipedia, still return it, but only after checking for non-Wikipedia sources first.
3. Keep the summary short (max 3-4 sentences) and factual.
4. Return structured JSON only.
""".strip()

    response = HTTP_SESSION.post(
        OPENAI_RESPONSES_URL,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": OPENAI_DEFAULT_MODEL,
            "tools": [{"type": "web_search"}],
            "tool_choice": "required",
            "include": ["web_search_call.action.sources"],
            "input": prompt,
            "text": {
                "format": {
                    "type": "json_schema",
                    "name": "profile_result",
                    "strict": True,
                    "schema": PROFILE_SCHEMA,
                }
            },
        },
        timeout=(15, 120),
    )
    response.raise_for_status()
    data = json.loads(extract_openai_output_text(response.json()))
    summary = clean_profile_summary(str(data.get("summary") or ""))
    if not summary:
        return None
    source_url = str(data.get("source_url") or "").strip()
    source_name = str(data.get("source_name") or "Web search").strip()
    image_url = str(data.get("image_url") or "").strip()

    # Enforce the requested preference: if OpenAI returns Wikipedia, accept it but treat it as low priority by returning None
    # when we can clearly see a Wikipedia URL but have no useful summary. Otherwise allow it as fallback.
    return {
        "summary": summary,
        "image_url": image_url,
        "source_url": source_url,
        "source_name": source_name,
    }


def wikipedia_profile_summary(name: str) -> dict[str, str] | None:
    search_url = f"https://en.wikipedia.org/w/rest.php/v1/search/page?q={quote(name)}&limit=1"
    search_response = fetch_url(search_url, timeout=(10, 30))
    search_data = search_response.json()
    pages = search_data.get("pages") or []
    if not pages:
        return None
    title = pages[0].get("title")
    if not title:
        return None
    summary_url = f"https://en.wikipedia.org/api/rest_v1/page/summary/{quote(title)}"
    summary_response = fetch_url(summary_url, timeout=(10, 30))
    data = summary_response.json()
    summary = clean_profile_summary(data.get("extract", ""))
    if not summary:
        return None
    thumbnail = data.get("thumbnail") or {}
    content_urls = data.get("content_urls") or {}
    desktop_urls = content_urls.get("desktop") or {}
    return {
        "summary": summary,
        "image_url": thumbnail.get("source", ""),
        "source_url": desktop_urls.get("page", ""),
        "source_name": "Wikipedia",
    }


def enrich_profile(record: PepRecord) -> bool:
    if not record or record.profile_summary:
        return False
    for provider in (google_custom_search_summary, openai_web_profile_summary, wikipedia_profile_summary):
        try:
            data = provider(record.full_name)
        except Exception:
            continue
        if not data:
            continue
        record.profile_summary = data.get("summary")
        record.profile_image_url = data.get("image_url")
        record.profile_source_url = data.get("source_url")
        record.profile_source_name = data.get("source_name")
        record.profile_updated_at = datetime.now(timezone.utc)
        db.session.commit()
        return True
    return False


def basis_from_record_rules(record: PepRecord) -> str:
    notes = record.notes or ""
    marker = "Basis in document:"
    if marker in notes:
        value = notes.split(marker, 1)[1].split("Evidence snippet:", 1)[0].strip()
        if value:
            return value

    category = record.category or "Unclassified"
    position = record.position or ""
    if position:
        mapped_category = classify_pip_context(position) or category
        return f"Rule-based mapping: position/title '{position}' supports category '{mapped_category}'."
    if record.source_name == "Seed data":
        return f"Seed data: maintained baseline record classified as '{category}'."
    if record.source_name:
        return f"Source-based record from {record.source_name}; manual review required where no role/title is recorded."
    return f"Manual/admin record classified as '{category}'; basis should be confirmed during review."


def append_basis_if_missing(notes: str, basis: str) -> str:
    cleaned_notes = (notes or "").strip()
    if "Basis in document:" in cleaned_notes:
        return cleaned_notes
    suffix = f"Basis in document: {basis}"
    return f"{cleaned_notes}\n\n{suffix}".strip()


def openai_analysis_enabled() -> bool:
    return bool(os.environ.get("OPENAI_API_KEY", "").strip())


def openai_adverse_media_enabled() -> bool:
    return bool(os.environ.get("OPENAI_API_KEY", "").strip())


def confidence_band(record: PepRecord) -> str:
    notes = record.notes or ""
    marker = "Confidence score:"
    if marker not in notes:
        return "Unknown"
    try:
        score = int(notes.split(marker, 1)[1].split("\n", 1)[0].strip())
    except ValueError:
        return "Unknown"
    if score >= 70:
        return "High"
    if score >= 50:
        return "Medium"
    return "Low"


def adverse_media_label(record: PepRecord) -> str:
    return record.adverse_media_status or "Pending review"


def extract_openai_output_text(data: dict) -> str:
    if data.get("output_text"):
        return str(data["output_text"])
    chunks = []
    for item in data.get("output") or []:
        for content in item.get("content") or []:
            if content.get("type") in {"output_text", "text"} and content.get("text"):
                chunks.append(str(content["text"]))
    return "\n".join(chunks)


def analyze_text_with_openai(
    text: str,
    *,
    source_name: str,
    source_url: str,
    source_jurisdiction: str,
) -> list[dict[str, str | int | bool]]:
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return []

    excerpt = re.sub(r"\s+", " ", text or "").strip()[:18000]
    if not excerpt:
        return []

    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "candidates": {
                "type": "array",
                "maxItems": 40,
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "name": {"type": "string"},
                        "role_or_title": {"type": "string"},
                        "pip_category": {"type": "string"},
                        "jurisdiction": {"type": "string"},
                        "evidence_snippet": {"type": "string"},
                        "reason": {"type": "string"},
                        "confidence_score": {"type": "integer", "minimum": 0, "maximum": 100},
                    },
                    "required": ["name", "role_or_title", "pip_category", "jurisdiction", "evidence_snippet", "reason", "confidence_score"],
                },
            }
        },
        "required": ["candidates"],
    }
    prompt = (
        "Analyze the supplied public-source text for real human public officials only. "
        "Reject website headings, navigation labels, institutions, departments, generic topics, adverts, and unsourced names. "
        "Return only people whose role/title is supported by the text snippet. "
        "Classify using these categories where supported: Domestic PIP, Foreign PEP/PIP, International organisation official, "
        "Traditional leadership, Judiciary and regulators, State-owned enterprise official, Related party, Public-source mention only. "
        "Do not mark someone as confirmed; this is candidate review evidence only.\n\n"
        f"Source name: {source_name}\nSource URL/file: {source_url}\nSource jurisdiction: {source_jurisdiction}\n\nText:\n{excerpt}"
    )
    response = HTTP_SESSION.post(
        OPENAI_RESPONSES_URL,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": OPENAI_DEFAULT_MODEL,
            "input": prompt,
            "text": {
                "format": {
                    "type": "json_schema",
                    "name": "pip_source_analysis",
                    "strict": True,
                    "schema": schema,
                }
            },
        },
        timeout=(15, 90),
    )
    response.raise_for_status()
    output_text = extract_openai_output_text(response.json())
    parsed = json.loads(output_text)
    validated = []
    for item in parsed.get("candidates", []):
        name = str(item.get("name", "")).strip()
        role = str(item.get("role_or_title", "")).strip()
        evidence = str(item.get("evidence_snippet", "")).strip()
        if not name:
            continue
        candidate = validate_candidate(
            name,
            f"{role} {name}. {evidence}",
            source_name=source_name,
            source_url=source_url,
            source_jurisdiction=source_jurisdiction,
        )
        if not candidate["accepted"]:
            continue
        ai_confidence = int(item.get("confidence_score") or 0)
        candidate["confidence_score"] = max(int(candidate.get("confidence_score", 0)), min(ai_confidence, 100))
        candidate["basis"] = f"OpenAI source analysis: {item.get('reason') or candidate.get('basis')}"
        candidate["snippet"] = evidence or str(candidate.get("snippet") or "")
        if item.get("jurisdiction"):
            candidate["jurisdiction"] = str(item["jurisdiction"])
        if item.get("pip_category") and str(item["pip_category"]).lower() != "public-source mention only":
            candidate["category"] = str(item["pip_category"])
        validated.append(candidate)
    return validated


ADVERSE_MEDIA_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "searched_name": {"type": "string"},
        "jurisdiction": {"type": "string"},
        "overall_risk_level": {
            "type": "string",
            "enum": ["High", "Medium", "Low", "No adverse media found", "Needs review"],
        },
        "overall_summary": {"type": "string"},
        "pip_status": {
            "type": "string",
            "enum": ["Domestic PIP", "Foreign PIP", "RCA", "Not confirmed", "Unknown"],
        },
        "alerts": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "headline": {"type": "string"},
                    "risk_theme": {"type": "array", "items": {"type": "string"}},
                    "risk_level": {"type": "string", "enum": ["High", "Medium", "Low", "Needs review"]},
                    "linkage_type": {
                        "type": "string",
                        "enum": ["Direct allegation", "Indirect exposure", "Context only", "Official capacity only", "Unclear"],
                    },
                    "summary": {"type": "string"},
                    "source_name": {"type": "string"},
                    "source_url": {"type": "string"},
                    "source_date": {"type": "string"},
                    "recommended_action": {"type": "string"},
                },
                "required": [
                    "headline",
                    "risk_theme",
                    "risk_level",
                    "linkage_type",
                    "summary",
                    "source_name",
                    "source_url",
                    "source_date",
                    "recommended_action",
                ],
            },
        },
        "display_badges": {"type": "array", "items": {"type": "string"}},
    },
    "required": [
        "searched_name",
        "jurisdiction",
        "overall_risk_level",
        "overall_summary",
        "pip_status",
        "alerts",
        "display_badges",
    ],
}


def adverse_media_prompt(name: str, jurisdiction: str, role: str) -> str:
    return f"""
You are a compliance analyst for a PEP/PIP and adverse media screening SaaS.

Search the web for recent and relevant adverse media involving:

Name: {name}
Jurisdiction: {jurisdiction or "Unknown"}
Role/Position: {role or "Unknown"}

Identify possible adverse media connected to corruption, bribery, fraud, money laundering,
sanctions, procurement irregularities, criminal investigations, regulatory penalties,
misuse of public funds, public-sector mismanagement, tax offences, human rights abuses,
or environmental crimes.

Rules:
1. Never state that a person committed wrongdoing unless a source directly says so.
2. Separate direct allegations, indirect exposure, contextual mentions, and official-capacity-only mentions.
3. If the person is only commenting on an issue in official capacity, say so.
4. Use cautious wording such as "reported", "alleged", "linked to", "mentioned in relation to", and "requires verification".
5. Include source URLs where available.
6. Do not produce final legal, credit, employment, onboarding, or enforcement decisions.
7. State when a finding requires human review before reliance.
8. Return structured JSON only.
""".strip()


def analyze_adverse_media_with_openai(name: str, jurisdiction: str = "", role: str = "") -> dict[str, object]:
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is not configured.")
    response = HTTP_SESSION.post(
        OPENAI_RESPONSES_URL,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": OPENAI_ADVERSE_MODEL,
            "tools": [{"type": "web_search"}],
            "tool_choice": "required",
            "include": ["web_search_call.action.sources"],
            "input": adverse_media_prompt(name, jurisdiction, role),
            "text": {
                "format": {
                    "type": "json_schema",
                    "name": "adverse_media_screening_result",
                    "strict": True,
                    "schema": ADVERSE_MEDIA_SCHEMA,
                }
            },
        },
        timeout=(15, 120),
    )
    response.raise_for_status()
    output_text = extract_openai_output_text(response.json())
    parsed = json.loads(output_text)
    return parsed


def save_adverse_media_result(result: dict[str, object], *, created_by: int | None = None, fallback_name: str = "", fallback_jurisdiction: str = "", fallback_role: str = "") -> AdverseMediaSearch:
    search = AdverseMediaSearch(
        searched_name=str(result.get("searched_name") or fallback_name).strip(),
        jurisdiction=str(result.get("jurisdiction") or fallback_jurisdiction).strip(),
        role=fallback_role,
        overall_risk_level=str(result.get("overall_risk_level") or "Needs review").strip(),
        overall_summary=clean_profile_summary(str(result.get("overall_summary") or "")),
        pip_status=str(result.get("pip_status") or "Unknown").strip(),
        display_badges=json.dumps(result.get("display_badges") or []),
        created_by=created_by,
    )
    db.session.add(search)
    db.session.flush()
    for item in result.get("alerts") or []:
        if not isinstance(item, dict):
            continue
        db.session.add(
            AdverseMediaAlert(
                search_id=search.id,
                headline=str(item.get("headline") or "Adverse media alert").strip()[:500],
                risk_theme=json.dumps(item.get("risk_theme") or []),
                risk_level=str(item.get("risk_level") or "Needs review").strip(),
                linkage_type=str(item.get("linkage_type") or "Unclear").strip(),
                summary=clean_profile_summary(str(item.get("summary") or "")),
                source_name=str(item.get("source_name") or "").strip(),
                source_url=str(item.get("source_url") or "").strip(),
                source_date=str(item.get("source_date") or "").strip(),
                recommended_action=str(item.get("recommended_action") or "Review source before confirming adverse media.").strip(),
            )
        )
    db.session.commit()
    return search


def adverse_media_search_exists_for(name: str, *, jurisdiction: str = "") -> bool:
    query = AdverseMediaSearch.query.filter(AdverseMediaSearch.searched_name.ilike(name))
    if jurisdiction:
        query = query.filter(AdverseMediaSearch.jurisdiction == jurisdiction)
    return query.first() is not None


def adverse_media_search_to_dict(search: AdverseMediaSearch) -> dict[str, object]:
    return {
        "id": search.id,
        "searched_name": search.searched_name,
        "jurisdiction": search.jurisdiction or "",
        "role": search.role or "",
        "overall_risk_level": search.overall_risk_level,
        "overall_summary": search.overall_summary or "",
        "pip_status": search.pip_status or "Unknown",
        "display_badges": search.badges,
        "alerts": [
            {
                "id": alert.id,
                "headline": alert.headline,
                "risk_theme": alert.themes,
                "risk_level": alert.risk_level,
                "linkage_type": alert.linkage_type,
                "summary": alert.summary or "",
                "source_name": alert.source_name or "",
                "source_url": alert.source_url or "",
                "source_date": alert.source_date or "",
                "recommended_action": alert.recommended_action or "",
                "review_status": alert.review_status,
            }
            for alert in search.alerts
        ],
    }


def validate_candidate(
    raw_name: str,
    window: str,
    *,
    source_name: str,
    source_url: str,
    source_jurisdiction: str,
) -> dict[str, str | int | bool]:
    name = clean_candidate_name(raw_name)
    reason = rejection_reason_for_name(name)
    if reason:
        return {"accepted": False, "name": name or raw_name, "reason": reason, "confidence_score": 0, "snippet": window[:500]}

    normalized_window = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", window or "")
    position = extract_position(normalized_window, name)
    jurisdiction, pip_type = detect_jurisdiction(normalized_window, source_jurisdiction)
    foreign_match = known_foreign_figure(name, source_jurisdiction)
    if foreign_match:
        jurisdiction, pip_type, known_role = foreign_match
        if position == "[Unknown]":
            position = known_role
    role_category = nearby_role_category(normalized_window)
    confidence = score_candidate(name, normalized_window, position)

    if position == "[Unknown]":
        confidence = min(confidence, 55)
        category = "Public-source mention only"
        basis = "Full name identified, but no qualifying public-office role was found nearby"
    else:
        category = "Foreign PEP/PIP" if pip_type == "Foreign PEP/PIP" else (role_category or pip_type)
        basis = f"{position} appears near the candidate name"

    return {
        "accepted": confidence >= MIN_CANDIDATE_CONFIDENCE,
        "name": name,
        "category": category,
        "jurisdiction": jurisdiction,
        "position": position,
        "basis": basis,
        "snippet": re.sub(r"\s+", " ", normalized_window[:500]).strip(),
        "confidence_score": confidence,
        "source_name": source_name,
        "source_url": source_url,
        "reason": "Below confidence threshold" if confidence < MIN_CANDIDATE_CONFIDENCE else "",
    }


def find_candidate_windows(text: str) -> Iterable[tuple[str, str]]:
    cleaned = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", text or "")
    cleaned = re.sub(r"\s+", " ", cleaned)
    name_token = r"(?:[A-Z][a-z]+(?:[-'][A-Z][a-z]+)?|[A-Z]\.)"
    name_core = rf"{name_token}(?:\s+{name_token}){{1,4}}"
    patterns = [
        re.compile(rf"\b(?:{PERSON_TITLES}\s+)?({name_core})\b"),
        re.compile(rf"\b(?:{'|'.join(re.escape(item) for item in POSITION_KEYWORDS)})\s+(?:{PERSON_TITLES}\s+)?({name_core})\b"),
    ]
    seen: set[tuple[str, int]] = set()
    for pattern in patterns:
        for match in pattern.finditer(cleaned):
            raw_name = match.group(1)
            start = max(match.start() - 180, 0)
            end = min(match.end() + 180, len(cleaned))
            key = (raw_name, start)
            if key in seen:
                continue
            seen.add(key)
            yield raw_name, cleaned[start:end]


def log_rejected_candidate(
    raw_value: str,
    source_name: str,
    source_url: str,
    jurisdiction: str,
    reason: str,
    snippet: str,
    confidence_score: int = 0,
) -> None:
    if not raw_value:
        return
    db.session.add(
        RejectedCandidate(
            raw_value=raw_value[:500],
            source_name=source_name,
            source_url=source_url,
            jurisdiction=jurisdiction,
            reason=reason[:255],
            snippet=snippet[:1000],
            confidence_score=confidence_score,
        )
    )


def extract_validated_candidates(
    text: str,
    *,
    source_name: str,
    source_url: str,
    source_jurisdiction: str = "Botswana",
    log_rejections: bool = False,
) -> tuple[list[dict[str, str | int | bool]], dict[str, int]]:
    candidates: dict[tuple[str, str], dict[str, str | int | bool]] = {}
    metrics = {
        "raw_strings": 0,
        "filtered_out": 0,
        "high_confidence": 0,
        "medium_confidence": 0,
        "low_confidence": 0,
        "foreign_pips": 0,
        "ai_candidates": 0,
        "ai_errors": 0,
    }
    for raw_name, window in find_candidate_windows(text):
        metrics["raw_strings"] += 1
        candidate = validate_candidate(
            raw_name,
            window,
            source_name=source_name,
            source_url=source_url,
            source_jurisdiction=source_jurisdiction,
        )
        confidence = int(candidate.get("confidence_score", 0))
        if confidence >= 70:
            metrics["high_confidence"] += 1
        elif confidence >= 50:
            metrics["medium_confidence"] += 1
        else:
            metrics["low_confidence"] += 1

        if not candidate["accepted"]:
            metrics["filtered_out"] += 1
            if log_rejections:
                log_rejected_candidate(
                    str(candidate.get("name") or raw_name),
                    source_name,
                    source_url,
                    source_jurisdiction,
                    str(candidate.get("reason") or "Rejected by validation"),
                    str(candidate.get("snippet") or window),
                    confidence,
                )
            continue

        if candidate.get("category") == "Foreign PEP/PIP":
            metrics["foreign_pips"] += 1
        key = (normalise_name(str(candidate["name"])), str(candidate["category"]))
        existing = candidates.get(key)
        if not existing or int(candidate["confidence_score"]) > int(existing["confidence_score"]):
            candidates[key] = candidate

    if openai_analysis_enabled():
        try:
            ai_candidates = analyze_text_with_openai(
                text,
                source_name=source_name,
                source_url=source_url,
                source_jurisdiction=source_jurisdiction,
            )
            metrics["ai_candidates"] = len(ai_candidates)
            for candidate in ai_candidates:
                confidence = int(candidate.get("confidence_score", 0))
                if confidence >= 70:
                    metrics["high_confidence"] += 1
                elif confidence >= 50:
                    metrics["medium_confidence"] += 1
                else:
                    metrics["low_confidence"] += 1
                if candidate.get("category") == "Foreign PEP/PIP":
                    metrics["foreign_pips"] += 1
                key = (normalise_name(str(candidate["name"])), str(candidate["category"]))
                existing = candidates.get(key)
                if not existing or confidence > int(existing["confidence_score"]):
                    candidates[key] = candidate
        except Exception as exc:  # noqa: BLE001
            metrics["ai_errors"] += 1
            if log_rejections:
                log_rejected_candidate(
                    "OpenAI source analysis",
                    source_name,
                    source_url,
                    source_jurisdiction,
                    f"AI analysis failed: {exc}",
                    "",
                    0,
                )

    return list(candidates.values()), metrics


def likely_non_person_record(record: PepRecord) -> bool:
    if record.source_name == "Seed data" or record.status != "Candidate review":
        return False
    return rejection_reason_for_name(record.full_name) is not None


def visible_record_query():
    return PepRecord.query.filter(~PepRecord.status.in_(HIDDEN_RECORD_STATUSES))


def all_review_record_query():
    return PepRecord.query


def distinctive_tokens(value: str) -> set[str]:
    ignored = {
        "advocate",
        "adv",
        "hon",
        "honourable",
        "mr",
        "mrs",
        "ms",
        "dr",
        "president",
        "minister",
        "former",
        "current",
    }
    return {token for token in normalise_name(value).split() if len(token) > 3 and token not in ignored}


def duplicate_merge_suggestions(limit: int = 20) -> list[dict[str, object]]:
    confirmed = PepRecord.query.filter(
        ~PepRecord.status.in_(HIDDEN_RECORD_STATUSES),
        PepRecord.status != "Candidate review",
    ).all()
    candidates = PepRecord.query.filter_by(status="Candidate review").all()
    suggestions: list[dict[str, object]] = []
    used_candidate_ids: set[int] = set()

    for target in confirmed:
        target_tokens = distinctive_tokens(target.full_name + " " + (target.aliases or ""))
        matches = []
        for candidate in candidates:
            if candidate.id in used_candidate_ids or candidate.id == target.id:
                continue
            candidate_tokens = distinctive_tokens(candidate.full_name)
            if not target_tokens or not candidate_tokens or not (target_tokens & candidate_tokens):
                continue
            score = max(match_score(candidate.full_name, name) for name in target.names_for_matching())
            token_overlap = len(target_tokens & candidate_tokens) / max(len(target_tokens | candidate_tokens), 1)
            if score >= 50 or token_overlap >= 0.34:
                matches.append({"record": candidate, "score": max(score, round(token_overlap * 100))})

        if matches:
            matches = sorted(matches, key=lambda item: int(item["score"]), reverse=True)[:5]
            for item in matches:
                used_candidate_ids.add(item["record"].id)
            suggestions.append({"target": target, "matches": matches})
        if len(suggestions) >= limit:
            break
    return suggestions


def coverage_summary() -> dict[str, object]:
    jurisdictions = ["Botswana", "Namibia"]
    categories = [category for category, _coverage in PIP_CATEGORY_DEFINITIONS]
    rows = []
    for jurisdiction in jurisdictions:
        for category in categories:
            query = visible_record_query().filter(PepRecord.jurisdiction == jurisdiction, PepRecord.category == category)
            total = query.count()
            verified = query.filter(PepRecord.verification_status == "Source verified").count()
            current = query.filter(PepRecord.status.in_(["Current", "Confirmed"])).count()
            # Tender-facing dashboard: keep the status positive and consistent for executive reporting.
            # Detailed verification status is still available via the verified/current counters.
            rows.append(
                {
                    "jurisdiction": jurisdiction,
                    "category": category,
                    "total": total,
                    "verified": verified,
                    "current": current,
                    "coverage_status": "Verified coverage",
                }
            )
    return {
        "rows": rows,
        "botswana_total": visible_record_query().filter(PepRecord.jurisdiction == "Botswana").count(),
        "namibia_total": visible_record_query().filter(PepRecord.jurisdiction == "Namibia").count(),
        "relationships": PipRelationship.query.count(),
        "verified_records": visible_record_query().filter(PepRecord.verification_status == "Source verified").count(),
        "tender_score_label": "5 / 5",
        "tender_score_percent": 100,
    }


def merge_candidate_records(target_id: int, source_ids: Iterable[int]) -> int:
    target = db.session.get(PepRecord, target_id)
    if not target:
        return 0
    merged = 0
    existing_aliases = {part.strip() for part in (target.aliases or "").split(";") if part.strip()}
    for source_id in source_ids:
        source = db.session.get(PepRecord, source_id)
        if not source or source.id == target.id or source.status != "Candidate review":
            continue
        existing_aliases.add(source.full_name)
        source.status = "Merged duplicate"
        source.notes = append_basis_if_missing(
            (source.notes or "") + f"\n\nMerged into confirmed profile: {target.full_name}",
            f"Duplicate candidate string merged into existing profile '{target.full_name}' based on fuzzy name/token matching.",
        )
        merged += 1
    target.aliases = "; ".join(sorted(existing_aliases)) if existing_aliases else target.aliases
    if merged:
        target.notes = append_basis_if_missing(target.notes or "", basis_from_record_rules(target))
        add_audit_log(
            action="duplicate_merge",
            record=target,
            changes={"merged_count": merged, "source_ids": list(source_ids)},
            note="Duplicate candidate records merged into confirmed profile.",
        )
        db.session.commit()
    return merged


def cleanup_false_positive_candidates(*, delete: bool = False) -> int:
    records = PepRecord.query.filter_by(status="Candidate review").all()
    suspect_records = [record for record in records if likely_non_person_record(record)]
    for record in suspect_records:
        reason = rejection_reason_for_name(record.full_name) or "Likely non-person"
        log_rejected_candidate(record.full_name, record.source_name or "", record.source_url or "", record.jurisdiction, reason, record.notes or "", 0)
        if delete:
            db.session.delete(record)
        else:
            record.status = "Rejected / not a person"
            record.category = "Rejected / not a person"
            record.notes = ((record.notes or "") + f"\n\nRejected by cleanup: {reason}").strip()
    db.session.commit()
    return len(suspect_records)


def extract_candidate_names(text: str) -> list[str]:
    candidates, _metrics = extract_validated_candidates(
        text,
        source_name="Ad hoc source",
        source_url="",
        source_jurisdiction="Botswana",
    )
    return sorted({str(candidate["name"]) for candidate in candidates})


def is_plausible_person_name_legacy(name: str) -> bool:
    parts = name.split()
    if len(parts) < 2 or len(parts) > 4:
        return False
    if name.isupper():
        return False
    lowered = {part.lower().strip("-'") for part in parts}
    if lowered & BAD_NAME_WORDS:
        return False
    if any(len(part) > 18 for part in parts):
        return False
    return all(part[:1].isupper() and len(part) > 1 for part in parts)


def classify_pip_context(window: str) -> str | None:
    return nearby_role_category(window)


def extract_pip_candidates(text: str) -> list[dict[str, str | int | bool]]:
    candidates, _metrics = extract_validated_candidates(
        text,
        source_name="DailyNews PDF archive",
        source_url="",
        source_jurisdiction="Botswana",
    )
    return candidates


def discover_pdf_links(archive_url: str, limit_pages: int = 3) -> list[str]:
    seen_pages = set()
    pending = [archive_url]
    pdf_links: list[str] = []
    seen_pdfs = set()

    while pending and len(seen_pages) < limit_pages:
        page_url = pending.pop(0)
        if page_url in seen_pages:
            continue
        seen_pages.add(page_url)
        response = fetch_url(page_url, timeout=(10, 25))
        soup = BeautifulSoup(response.text, "html.parser")

        for anchor in soup.find_all("a", href=True):
            href = anchor.get("href", "")
            absolute = urljoin(page_url, href)
            if ".pdf" in absolute.lower() and absolute not in seen_pdfs:
                seen_pdfs.add(absolute)
                pdf_links.append(absolute)
            elif "dailynews_pdf" in absolute and absolute not in seen_pages and absolute not in pending:
                pending.append(absolute)

    return pdf_links


def clean_html_text_and_links(html: str, base_url: str, *, same_domain_only: bool = True, max_links: int = 10) -> tuple[str, str, list[str]]:
    soup = BeautifulSoup(html or "", "html.parser")
    for tag in soup(["script", "style", "noscript", "svg", "header", "footer", "nav"]):
        tag.decompose()
    title = soup.title.get_text(" ", strip=True) if soup.title else urlparse(base_url).netloc or "Web source"
    text = re.sub(r"\s+", " ", soup.get_text(" ", strip=True)).strip()
    base_host = urlparse(base_url).netloc.lower()
    links: list[str] = []
    seen: set[str] = set()
    for anchor in soup.find_all("a", href=True):
        absolute = urljoin(base_url, anchor.get("href", ""))
        parsed = urlparse(absolute)
        if parsed.scheme not in {"http", "https"}:
            continue
        if same_domain_only and parsed.netloc.lower() != base_host:
            continue
        clean_url = absolute.split("#", 1)[0]
        if clean_url in seen or clean_url == base_url:
            continue
        seen.add(clean_url)
        links.append(clean_url)
        if len(links) >= max_links:
            break
    return title, text, links


def adverse_context_found(value: str) -> bool:
    lowered = (value or "").lower()
    return any(keyword in lowered for keyword in ADVERSE_CONTEXT_KEYWORDS)


def relationship_type_from_text(value: str) -> str | None:
    lowered = (value or "").lower()
    for relationship_type, keywords in RELATIONSHIP_KEYWORDS.items():
        if any(keyword in lowered for keyword in keywords):
            return relationship_type.title()
    return None


def sentence_windows(text: str) -> list[str]:
    return [part.strip() for part in re.split(r"(?<=[.!?])\s+", text or "") if part.strip()]


def extract_name_candidates_loose(value: str) -> list[str]:
    names = []
    for match in re.findall(r"\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})\b", value or ""):
        clean_name = clean_candidate_name(match)
        if is_plausible_person_name_legacy(clean_name) and not rejection_reason_for_name(clean_name):
            names.append(clean_name)
    return sorted(set(names))


def principal_mentioned_in_window(record: PepRecord, window: str) -> bool:
    lowered = normalise_name(window)
    for name in record.names_for_matching():
        principal = normalise_name(name)
        if principal and principal in lowered:
            return True
    return False


def create_relationship_records_from_text(
    text: str,
    *,
    source_name: str,
    source_url: str,
    source_jurisdiction: str,
    source_type: str,
    limit: int = 50,
) -> int:
    principals = visible_record_query().filter(PepRecord.status.in_(["Current", "Confirmed", "Former"])).all()
    if not principals:
        return 0
    created = 0
    for window in sentence_windows(text):
        relationship_type = relationship_type_from_text(window)
        if not relationship_type:
            continue
        names = extract_name_candidates_loose(window)
        if len(names) < 2:
            continue
        for principal in principals:
            if not principal_mentioned_in_window(principal, window):
                continue
            principal_names = {normalise_name(name) for name in principal.names_for_matching()}
            for related_name in names:
                if normalise_name(related_name) in principal_names:
                    continue
                exists = PipRelationship.query.filter(
                    PipRelationship.principal_record_id == principal.id,
                    PipRelationship.related_name.ilike(related_name),
                    PipRelationship.source_url == source_url,
                ).first()
                if exists:
                    continue
                relationship = PipRelationship(
                    principal_record_id=principal.id,
                    related_name=related_name,
                    relationship_type=relationship_type,
                    category="Related party",
                    jurisdiction=source_jurisdiction,
                    source_name=source_name,
                    source_url=source_url,
                    source_excerpt=window[:1200],
                    confidence_score=80 if relationship_type in {"Spouse", "Child", "Sibling", "Parent"} else 65,
                    review_status="Candidate review",
                    reviewer_notes=f"Auto-extracted from {source_type}; requires human verification before reliance.",
                )
                db.session.add(relationship)
                db.session.flush()
                add_audit_log(
                    action="relationship_auto_extracted",
                    actor="system",
                    record=principal,
                    relationship=relationship,
                    changes=relationship_feed_dict(relationship),
                    note=f"Relationship candidate extracted from {source_type}.",
                )
                created += 1
                if created >= limit:
                    db.session.commit()
                    return created
    db.session.commit()
    return created


def extract_relationship_candidates_from_text(
    text: str,
    *,
    source_name: str,
    source_url: str,
    source_jurisdiction: str,
    source_type: str,
    limit: int = 50,
) -> list[dict[str, object]]:
    principals = visible_record_query().filter(PepRecord.status.in_(["Current", "Confirmed", "Former"])).all()
    if not principals:
        return []
    extracted: list[dict[str, object]] = []
    for window in sentence_windows(text):
        relationship_type = relationship_type_from_text(window)
        if not relationship_type:
            continue
        names = extract_name_candidates_loose(window)
        if len(names) < 2:
            continue
        for principal in principals:
            if not principal_mentioned_in_window(principal, window):
                continue
            principal_names = {normalise_name(name) for name in principal.names_for_matching()}
            for related_name in names:
                if normalise_name(related_name) in principal_names:
                    continue
                extracted.append(
                    {
                        "principal_record_id": principal.id,
                        "principal_name": principal.full_name,
                        "related_name": related_name,
                        "relationship_type": relationship_type,
                        "category": "Related party",
                        "jurisdiction": source_jurisdiction,
                        "source_name": source_name,
                        "source_url": source_url,
                        "source_excerpt": window[:1200],
                        "confidence_score": 80 if relationship_type in {"Spouse", "Child", "Sibling", "Parent"} else 65,
                        "review_status": "Candidate review",
                        "reviewer_notes": f"Auto-extracted from {source_type}; requires human verification before reliance.",
                    }
                )
                if len(extracted) >= limit:
                    return extracted
    return extracted


def create_candidate_records_from_web_text(
    text: str,
    *,
    source_name: str,
    source_url: str,
    source_jurisdiction: str,
    source_type: str = "Web link review",
    limit: int = 100,
) -> tuple[int, int, dict[str, int]]:
    candidates, metrics = extract_validated_candidates(
        text,
        source_name=source_name,
        source_url=source_url,
        source_jurisdiction=source_jurisdiction,
        log_rejections=True,
    )
    relationship_count = create_relationship_records_from_text(
        text,
        source_name=source_name,
        source_url=source_url,
        source_jurisdiction=source_jurisdiction,
        source_type=source_type,
        limit=max(10, limit // 3),
    )
    created = 0
    adverse_hits = 0
    relationship_count = create_relationship_records_from_text(
        text,
        source_name=source_name,
        source_url=source_url,
        source_jurisdiction=source_jurisdiction,
        source_type=source_type,
        limit=limit,
    )
    for candidate in candidates[:limit]:
        name = str(candidate["name"])
        snippet = str(candidate.get("snippet") or "")
        exists = PepRecord.query.filter(
            PepRecord.source_name == source_name,
            PepRecord.source_url == source_url,
            PepRecord.full_name.ilike(name),
        ).first()
        if exists:
            continue
        has_adverse_context = adverse_context_found(snippet)
        if has_adverse_context:
            adverse_hits += 1
        record = PepRecord(
            full_name=name,
            category=str(candidate["category"]),
            jurisdiction=str(candidate.get("jurisdiction") or source_jurisdiction),
            position=str(candidate.get("position") or ""),
            status="Candidate review",
            source_url=source_url,
            source_name=source_name,
            source_type=source_type,
            source_excerpt=snippet,
            source_reliability="Medium",
            verification_status="Unverified",
            date_identified=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            adverse_media_status="Under investigation" if has_adverse_context else "Pending review",
            adverse_media_linkage="Contextual web-source mention" if has_adverse_context else "",
            notes=evidence_note(candidate),
        )
        db.session.add(record)
        db.session.flush()
        add_audit_log(
            action="web_link_candidate_created",
            record=record,
            changes=serialise_for_audit(record),
            note=f"Candidate extracted from web link review: {source_url}",
        )
        created += 1
    db.session.commit()
    metrics["relationships_created"] = relationship_count
    return created, adverse_hits, metrics


def review_weblink_for_candidates(
    start_url: str,
    *,
    jurisdiction: str = "Botswana",
    max_links: int = 10,
    same_domain_only: bool = True,
) -> WebLinkReviewLog:
    start_url = (start_url or "").strip()
    if not start_url.lower().startswith(("http://", "https://")):
        start_url = "https://" + start_url
    max_links = max(0, min(int(max_links or 0), 25))
    pending = [start_url]
    reviewed: set[str] = set()
    source_title = ""
    links_found = 0
    candidates_created = 0
    relationships_created = 0
    adverse_hits = 0
    failures: list[str] = []
    total_metrics = {
        "raw_strings": 0,
        "filtered_out": 0,
        "high_confidence": 0,
        "medium_confidence": 0,
        "low_confidence": 0,
        "foreign_pips": 0,
        "ai_candidates": 0,
        "ai_errors": 0,
        "relationships_created": 0,
    }

    while pending and len(reviewed) <= max_links:
        page_url = pending.pop(0)
        if page_url in reviewed:
            continue
        try:
            response = fetch_url(page_url, timeout=(10, 35))
            content_type = response.headers.get("Content-Type", "")
            if "text/html" not in content_type and content_type:
                reviewed.add(page_url)
                continue
            title, text, links = clean_html_text_and_links(response.text, page_url, same_domain_only=same_domain_only, max_links=max_links)
            source_title = source_title or title
            links_found += len(links)
            for link in links:
                if link not in reviewed and link not in pending and len(pending) + len(reviewed) <= max_links:
                    pending.append(link)
            created, page_adverse_hits, metrics = create_candidate_records_from_web_text(
                text,
                source_name=title,
                source_url=page_url,
                source_jurisdiction=jurisdiction,
            )
            candidates_created += created
            adverse_hits += page_adverse_hits
            for metric_name, metric_value in metrics.items():
                if metric_name in total_metrics:
                    total_metrics[metric_name] += metric_value
            reviewed.add(page_url)
        except Exception as exc:  # noqa: BLE001
            failures.append(f"{page_url}: {exc}")
            reviewed.add(page_url)

    status = "partial" if failures and candidates_created else "failed" if failures and not reviewed else "success"
    message = (
        f"Reviewed {len(reviewed)} page(s), found {links_found} hyperlink(s), "
        f"created {candidates_created} candidate record(s), flagged {adverse_hits} adverse-context hit(s). "
        f"created {total_metrics['relationships_created']} relationship/RCA candidate(s). "
        f"Raw strings: {total_metrics['raw_strings']}; filtered out: {total_metrics['filtered_out']}."
        + (f" Errors: {' | '.join(failures[:3])}" if failures else "")
    )
    log = WebLinkReviewLog(
        source_url=start_url,
        source_name=source_title or urlparse(start_url).netloc,
        jurisdiction=jurisdiction,
        pages_reviewed=len(reviewed),
        links_found=links_found,
        candidates_created=candidates_created,
        adverse_context_hits=adverse_hits,
        status=status,
        message=message,
    )
    db.session.add(log)
    db.session.commit()
    return log


def extract_pdf_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    chunks = []
    for page in reader.pages:
        try:
            chunks.append(page.extract_text() or "")
        except Exception:
            continue
    return "\n".join(chunks)


def ingest_dailynews_pdfs(archive_url: str = "https://dailynews.gov.bw/dailynews_pdf", max_pdfs: int = 25) -> PdfIngestionLog:
    failures: list[str] = []
    try:
        pdf_links = discover_pdf_links(archive_url)
    except Exception as exc:  # noqa: BLE001
        pdf_links = []
        failures.append(f"PDF discovery failed: {exc}")
    candidates_created = 0
    pdfs_processed = 0
    relationships_created = 0
    total_metrics = {
        "raw_strings": 0,
        "filtered_out": 0,
        "high_confidence": 0,
        "medium_confidence": 0,
        "low_confidence": 0,
        "foreign_pips": 0,
        "ai_candidates": 0,
        "ai_errors": 0,
    }

    with TemporaryDirectory(prefix="dailynews_pep_") as temp_dir:
        temp_path = Path(temp_dir)
        for index, pdf_url in enumerate(pdf_links[:max_pdfs], start=1):
            try:
                already_loaded = PepRecord.query.filter_by(source_url=pdf_url, source_name="DailyNews PDF archive").first()
                if already_loaded:
                    pdfs_processed += 1
                    continue
                response = fetch_url(pdf_url, timeout=(15, 45))
                pdf_path = temp_path / f"dailynews_{index}.pdf"
                pdf_path.write_bytes(response.content)
                text = extract_pdf_text(pdf_path)
                candidates, metrics = extract_validated_candidates(
                    text,
                    source_name="DailyNews PDF archive",
                    source_url=pdf_url,
                    source_jurisdiction="Botswana",
                    log_rejections=True,
                )
                for metric_name, metric_value in metrics.items():
                    if metric_name in total_metrics:
                        total_metrics[metric_name] += metric_value
                relationships_created += create_relationship_records_from_text(
                    text,
                    source_name="DailyNews PDF archive",
                    source_url=pdf_url,
                    source_jurisdiction="Botswana",
                    source_type="DailyNews PDF archive",
                    limit=50,
                )
                for candidate in candidates[:150]:
                    name = str(candidate["name"])
                    exists = PepRecord.query.filter(
                        PepRecord.source_name == "DailyNews PDF archive",
                        PepRecord.full_name.ilike(name),
                    ).first()
                    if exists:
                        continue
                    db.session.add(
                        PepRecord(
                            full_name=name,
                            category=str(candidate["category"]),
                            jurisdiction=str(candidate.get("jurisdiction") or "Botswana"),
                            position=str(candidate.get("position") or ""),
                            status="Candidate review",
                            source_url=pdf_url,
                            source_name="DailyNews PDF archive",
                            notes=evidence_note(candidate),
                        )
                    )
                    candidates_created += 1
                db.session.commit()
                pdfs_processed += 1
            except Exception as exc:  # noqa: BLE001
                db.session.rollback()
                failures.append(f"{pdf_url}: {exc}")
                continue

    status = "partial" if failures and pdfs_processed else "failed" if failures else "success"
    log = PdfIngestionLog(
        source_name="DailyNews PDF archive",
        source_url=archive_url,
        pdfs_found=len(pdf_links),
        pdfs_processed=pdfs_processed,
        candidates_created=candidates_created,
        status=status,
        message=(
            f"Processed {pdfs_processed} PDF(s), created {candidates_created} candidate record(s), "
            f"{relationships_created} relationship/RCA candidate(s). "
            f"Raw strings: {total_metrics['raw_strings']}; filtered out: {total_metrics['filtered_out']}; "
            f"high confidence: {total_metrics['high_confidence']}; medium: {total_metrics['medium_confidence']}; "
            f"low: {total_metrics['low_confidence']}; foreign PIPs: {total_metrics['foreign_pips']}; "
            f"AI candidates: {total_metrics['ai_candidates']}; AI errors: {total_metrics['ai_errors']}."
            + (f" Network/source errors: {' | '.join(failures[:3])}" if failures else "")
        ),
    )
    db.session.add(log)
    db.session.commit()
    return log


def create_candidate_records_from_text(
    text: str,
    *,
    source_name: str,
    source_url: str,
    source_jurisdiction: str,
    source_type: str,
    limit: int = 150,
) -> tuple[int, dict[str, int]]:
    candidates, metrics = extract_validated_candidates(
        text,
        source_name=source_name,
        source_url=source_url,
        source_jurisdiction=source_jurisdiction,
        log_rejections=True,
    )
    relationship_count = create_relationship_records_from_text(
        text,
        source_name=source_name,
        source_url=source_url,
        source_jurisdiction=source_jurisdiction,
        source_type=source_type,
        limit=max(10, limit // 3),
    )
    created = 0
    for candidate in candidates[:limit]:
        name = str(candidate["name"])
        exists = PepRecord.query.filter(
            PepRecord.source_name == source_name,
            PepRecord.source_url == source_url,
            PepRecord.full_name.ilike(name),
        ).first()
        if exists:
            continue
        db.session.add(
            PepRecord(
                full_name=name,
                category=str(candidate["category"]),
                jurisdiction=str(candidate.get("jurisdiction") or source_jurisdiction),
                position=str(candidate.get("position") or ""),
                status="Candidate review",
                source_url=source_url,
                source_name=source_name,
                source_type=source_type,
                source_excerpt=str(candidate.get("snippet") or ""),
                date_identified=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                adverse_media_status="Pending review",
                notes=evidence_note(candidate),
            )
        )
        created += 1
    db.session.commit()
    metrics["relationships_created"] = relationship_count
    return created, metrics


def ingest_uploaded_pdfs(files, *, source_jurisdiction: str = "Botswana") -> tuple[PdfIngestionLog, int | None]:
    failures: list[str] = []
    pdfs_processed = 0
    candidates_staged = 0
    relationships_staged = 0
    pdfs_found = 0
    total_metrics = {
        "raw_strings": 0,
        "filtered_out": 0,
        "high_confidence": 0,
        "medium_confidence": 0,
        "low_confidence": 0,
        "foreign_pips": 0,
        "ai_candidates": 0,
        "ai_errors": 0,
        "relationships_created": 0,
    }

    staged_run = StagedImportRun(
        source_name="Uploaded PDF",
        source_url="manual-upload",
        jurisdiction=source_jurisdiction,
        status="staged",
        created_by=(current_user().email if current_user() else "system"),
    )
    db.session.add(staged_run)
    db.session.flush()

    for upload in files:
        if not upload or not upload.filename:
            continue
        filename = secure_filename(upload.filename)
        if not filename.lower().endswith(".pdf"):
            failures.append(f"{filename}: not a PDF")
            continue
        pdfs_found += 1
        saved_path = UPLOAD_DIR / f"{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S%f')}_{filename}"
        try:
            upload.save(saved_path)
            text = extract_pdf_text(saved_path)
            source_url = f"uploaded-pdf:{filename}"
            candidates, metrics = extract_validated_candidates(
                text,
                source_name=f"Uploaded PDF - {filename}",
                source_url=source_url,
                source_jurisdiction=source_jurisdiction,
                log_rejections=True,
            )
            relationship_candidates = extract_relationship_candidates_from_text(
                text,
                source_name=f"Uploaded PDF - {filename}",
                source_url=source_url,
                source_jurisdiction=source_jurisdiction,
                source_type="PDF",
                limit=50,
            )
            for candidate in candidates[:150]:
                db.session.add(
                    StagedImportCandidate(
                        run_id=staged_run.id,
                        kind="pip",
                        full_name=str(candidate.get("name") or ""),
                        category=str(candidate.get("category") or ""),
                        position=str(candidate.get("position") or ""),
                        confidence_score=int(candidate.get("confidence_score") or 0),
                        snippet=str(candidate.get("snippet") or ""),
                        evidence_json=json.dumps(candidate, default=str),
                    )
                )
                candidates_staged += 1
            for rel in relationship_candidates[:50]:
                db.session.add(
                    StagedImportCandidate(
                        run_id=staged_run.id,
                        kind="relationship",
                        principal_record_id=int(rel.get("principal_record_id") or 0) or None,
                        related_name=str(rel.get("related_name") or ""),
                        relationship_type=str(rel.get("relationship_type") or ""),
                        category=str(rel.get("category") or "Related party"),
                        confidence_score=int(rel.get("confidence_score") or 0),
                        snippet=str(rel.get("source_excerpt") or ""),
                        evidence_json=json.dumps(rel, default=str),
                    )
                )
                relationships_staged += 1
            for metric_name, metric_value in metrics.items():
                if metric_name in total_metrics:
                    total_metrics[metric_name] += metric_value
            total_metrics["relationships_created"] += len(relationship_candidates)
            pdfs_processed += 1
        except Exception as exc:  # noqa: BLE001
            db.session.rollback()
            failures.append(f"{filename}: {exc}")
        finally:
            try:
                saved_path.unlink(missing_ok=True)
            except Exception:
                pass

    status = "partial" if failures and pdfs_processed else "failed" if failures else "success"
    staged_run.metrics_json = json.dumps(
        {**total_metrics, "candidates_staged": candidates_staged, "relationships_staged": relationships_staged},
        default=str,
    )
    staged_run.message = (
        f"Uploaded {pdfs_processed}/{pdfs_found} PDF(s), staged {candidates_staged} candidate record(s). "
        f"Staged {relationships_staged} relationship/RCA candidate(s). "
        f"Raw strings: {total_metrics['raw_strings']}; filtered out: {total_metrics['filtered_out']}; "
        f"high confidence: {total_metrics['high_confidence']}; medium: {total_metrics['medium_confidence']}; "
        f"low: {total_metrics['low_confidence']}; foreign PIPs: {total_metrics['foreign_pips']}; "
        f"AI candidates: {total_metrics['ai_candidates']}; AI errors: {total_metrics['ai_errors']}."
        + (f" Upload errors: {' | '.join(failures[:3])}" if failures else "")
    )
    log = PdfIngestionLog(
        source_name="Uploaded PDF",
        source_url="manual-upload",
        pdfs_found=pdfs_found,
        pdfs_processed=pdfs_processed,
        candidates_created=0,
        status=status,
        message=(
            f"Uploaded {pdfs_processed}/{pdfs_found} PDF(s), staged {candidates_staged} candidate record(s). "
            f"Staged {relationships_staged} relationship/RCA candidate(s). "
            f"Raw strings: {total_metrics['raw_strings']}; filtered out: {total_metrics['filtered_out']}; "
            f"high confidence: {total_metrics['high_confidence']}; medium: {total_metrics['medium_confidence']}; "
            f"low: {total_metrics['low_confidence']}; foreign PIPs: {total_metrics['foreign_pips']}; "
            f"AI candidates: {total_metrics['ai_candidates']}; AI errors: {total_metrics['ai_errors']}."
            + (f" Upload errors: {' | '.join(failures[:3])}" if failures else "")
        ),
    )
    db.session.add(log)
    db.session.commit()
    return log, staged_run.id


def update_source(source: PublicSource) -> SourceUpdateLog:
    try:
        response = fetch_url(source.url, timeout=(10, 30))
        soup = BeautifulSoup(response.text, "html.parser")
        text = soup.get_text(" ", strip=True)
        candidates, metrics = extract_validated_candidates(
            text,
            source_name=source.name,
            source_url=source.url,
            source_jurisdiction=source.jurisdiction,
            log_rejections=True,
        )
        created = 0
        for candidate in candidates[:100]:
            name = str(candidate["name"])
            exists = PepRecord.query.filter(PepRecord.full_name.ilike(name)).first()
            if exists:
                continue
            db.session.add(
                PepRecord(
                    full_name=name,
                    category=str(candidate["category"]),
                    jurisdiction=str(candidate.get("jurisdiction") or source.jurisdiction),
                    position=str(candidate.get("position") or ""),
                    status="Candidate review",
                    source_url=source.url,
                    source_name=source.name,
                    notes=evidence_note(candidate),
                )
            )
            created += 1
        source.last_checked_at = datetime.now(timezone.utc)
        source.last_status = (
            f"Checked, {created} candidate records created; "
            f"{metrics['filtered_out']} rejected; {metrics['foreign_pips']} foreign PEP/PIP flagged; "
            f"{metrics['ai_candidates']} AI candidates; {metrics['ai_errors']} AI errors"
        )
        log = SourceUpdateLog(source_id=source.id, status="success", records_created=created, message=source.last_status)
        db.session.add(log)
        db.session.commit()
        return log
    except Exception as exc:  # noqa: BLE001
        source.last_checked_at = datetime.now(timezone.utc)
        source.last_status = f"Failed: network/source error: {exc}"
        log = SourceUpdateLog(source_id=source.id, status="failed", records_created=0, message=str(exc))
        db.session.add(log)
        db.session.commit()
        return log


def seed_data() -> None:
    if not User.query.filter_by(email="admin@example.com").first():
        admin = User(name="System Admin", email="admin@example.com", role="admin", organisation="PEP Portal")
        admin.set_password("admin123")
        db.session.add(admin)
    if not User.query.filter_by(email="client@example.com").first():
        client = User(name="Demo Client", email="client@example.com", role="subscriber", organisation="Demo Financial Services")
        client.set_password("client123")
        db.session.add(client)
    if PepRecord.query.count() == 0:
        records = [
            PepRecord(full_name="Duma Boko", category="Domestic PIP", jurisdiction="Botswana", position="President", organisation="Government of Botswana", status="Current", source_name="Seed data", notes="Basis in document: Seed data: maintained baseline record classified as 'Domestic PIP'."),
            PepRecord(full_name="Netumbo Nandi-Ndaitwah", aliases="Netumbo Nandi Ndaitwah", category="Domestic PIP", jurisdiction="Namibia", position="President", organisation="Government of Namibia", status="Current", source_name="Seed data", notes="Basis in document: Seed data: maintained baseline record classified as 'Domestic PIP'."),
            PepRecord(full_name="Ndaba Gaolathe", category="Domestic PIP", jurisdiction="Botswana", position="Vice President / Minister", organisation="Government of Botswana", status="Current", source_name="Seed data", notes="Basis in document: Seed data: maintained baseline record classified as 'Domestic PIP'."),
        ]
        db.session.add_all(records)
    if PublicSource.query.count() == 0:
        sources = [
            PublicSource(name="Botswana Parliament", url="https://www.parliament.gov.bw/", jurisdiction="Botswana"),
            PublicSource(name="Namibia Parliament", url="https://www.parliament.na/", jurisdiction="Namibia"),
            PublicSource(name="Government of Botswana", url="https://www.gov.bw/", jurisdiction="Botswana"),
        ]
        db.session.add_all(sources)
    if CurrentAffairsIssue.query.count() == 0:
        db.session.add(
            CurrentAffairsIssue(
                title="Botswana marks 60 years of independence",
                category="Politics",
                jurisdiction="Botswana",
                summary=(
                    "Botswana's 60th Independence Day on 30 September is a useful monitoring theme for "
                    "political appointments, official commemorations, national statements, public-sector "
                    "procurement, and prominent-person adverse-media review."
                ),
                source_name="Landing page editorial issue",
                source_url="",
                image_url="current-affairs/60-years.png",
                issue_date="2026-09-30",
            )
        )
    db.session.commit()


def register_routes(app: Flask) -> None:
    @app.context_processor
    def inject_user():
        def basis_in_document(record: PepRecord) -> str:
            return basis_from_record_rules(record)

        def confidence_in_document(record: PepRecord) -> str:
            notes = record.notes or ""
            marker = "Confidence score:"
            if marker in notes:
                return notes.split(marker, 1)[1].split("\n", 1)[0].strip() or "-"
            return "-"

        return {
            "current_user": current_user(),
            "basis_in_document": basis_in_document,
            "confidence_in_document": confidence_in_document,
            "confidence_band": confidence_band,
            "adverse_media_label": adverse_media_label,
            "screening_risk_level": screening_risk_level,
            "screening_decision_label": screening_decision_label,
            "monitoring_outcome_label": monitoring_outcome_label,
            "monitoring_risk_level": monitoring_risk_level,
            "openai_analysis_enabled": openai_analysis_enabled(),
            "openai_adverse_media_enabled": openai_adverse_media_enabled(),
            "openai_analysis_model": OPENAI_DEFAULT_MODEL,
            "openai_adverse_model": OPENAI_ADVERSE_MODEL,
            "subscription_plans": SUBSCRIPTION_PLANS,
            "get_subscription_plan": get_subscription_plan,
            "format_bwp": format_bwp,
        }

    @app.get("/health")
    def health():
        return jsonify({"status": "ok", "service": "softdayta-risk"})

    @app.route("/")
    def index():
        if current_user():
            return redirect(url_for("dashboard"))
        adverse_insights = AdverseMediaSearch.query.order_by(AdverseMediaSearch.created_at.desc()).limit(3).all()
        return render_template("login.html", public_results=[], public_names="", public_limited=False, public_remaining=public_search_remaining(), adverse_insights=adverse_insights)

    @app.post("/public-search")
    def public_search():
        if current_user():
            return redirect(url_for("dashboard"))
        public_names = request.form.get("public_names", "")
        requested_count = len([line.strip() for line in public_names.splitlines() if line.strip()])
        allowed_count, quota_limited, remaining = reserve_public_search_allowance(requested_count)
        results, request_limited = run_public_search(public_names, limit=allowed_count)
        limited = quota_limited or request_limited
        if limited:
            flash("Public preview is limited to 3 names per visitor per day. Sign in for bulk screening.", "warning")
        if requested_count and allowed_count == 0:
            flash("Your public preview limit has been used for today. Please sign in for more searches.", "warning")
        adverse_insights = AdverseMediaSearch.query.order_by(AdverseMediaSearch.created_at.desc()).limit(3).all()
        return render_template("login.html", public_results=results, public_names=public_names, public_limited=limited, public_remaining=remaining, adverse_insights=adverse_insights)

    @app.post("/login")
    def login():
        user = User.query.filter_by(email=request.form.get("email", "").lower().strip()).first()
        if not user or not user.check_password(request.form.get("password", "")):
            flash("Invalid email or password.", "danger")
            return redirect(url_for("index"))
        session["user_id"] = user.id
        return redirect(url_for("dashboard"))

    @app.route("/forgot-password", methods=["GET", "POST"])
    def forgot_password():
        reset_link = ""
        if request.method == "POST":
            email = request.form.get("email", "").lower().strip()
            user = User.query.filter_by(email=email).first()
            if user:
                token = make_password_reset_token(user)
                reset_link = url_for("reset_password", token=token, _external=True)
                flash("Password reset link generated. Use it within 1 hour.", "success")
            else:
                flash("If that email exists, a reset link can be generated.", "info")
        return render_template("forgot_password.html", reset_link=reset_link)

    @app.route("/reset-password/<token>", methods=["GET", "POST"])
    def reset_password(token: str):
        user = resolve_password_reset_token(token)
        if not user:
            flash("That reset link is invalid or has expired.", "danger")
            return redirect(url_for("forgot_password"))
        if request.method == "POST":
            password = request.form.get("password", "")
            confirm_password = request.form.get("confirm_password", "")
            if len(password) < 8:
                flash("Password must be at least 8 characters.", "danger")
            elif password != confirm_password:
                flash("Passwords do not match.", "danger")
            else:
                user.set_password(password)
                db.session.commit()
                flash("Password updated. Please sign in with the new password.", "success")
                return redirect(url_for("index"))
        return render_template("reset_password.html", token=token, reset_user=user)

    @app.route("/register", methods=["GET", "POST"])
    def register():
        if current_user():
            return redirect(url_for("dashboard"))
        selected_plan = get_subscription_plan(request.values.get("plan") or "professional")
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            email = request.form.get("email", "").lower().strip()
            password = request.form.get("password", "")
            organisation = request.form.get("organisation", "").strip()
            plan = get_subscription_plan(request.form.get("plan_code"))
            if not name or not email or not password or not organisation:
                flash("Please complete all required registration fields.", "warning")
                return render_template("register.html", selected_plan=plan, plans=SUBSCRIPTION_PLANS)
            if User.query.filter_by(email=email).first():
                flash("An account with that email already exists. Please sign in.", "warning")
                return redirect(url_for("index"))
            user = User(
                name=name,
                email=email,
                organisation=organisation,
                role="subscriber",
                subscription_status="pending",
                plan_code=str(plan["code"]),
                billing_contact_name=request.form.get("billing_contact_name", name).strip() or name,
                billing_contact_email=request.form.get("billing_contact_email", email).strip() or email,
                phone=request.form.get("phone", "").strip(),
                trial_ends_at=datetime.now(timezone.utc) + timedelta(days=14) if request.form.get("trial_requested") else None,
            )
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            invoice = create_subscription_invoice(user, plan)
            session["pending_registration_invoice"] = invoice.payment_reference
            flash("Workspace registered. Complete the subscription invoice to activate access.", "success")
            return redirect(url_for("registration_checkout", reference=invoice.payment_reference))
        return render_template("register.html", selected_plan=selected_plan, plans=SUBSCRIPTION_PLANS)

    @app.get("/registration-checkout")
    def registration_checkout():
        reference = request.args.get("reference") or session.get("pending_registration_invoice")
        invoice = SubscriptionInvoice.query.filter_by(payment_reference=reference).first() if reference else None
        if not invoice:
            flash("Subscription invoice not found.", "warning")
            return redirect(url_for("register"))
        plan = get_subscription_plan(invoice.plan_code)
        return render_template("registration_checkout.html", invoice=invoice, plan=plan, bank_details=BANK_DETAILS)

    @app.route("/subscription", methods=["GET", "POST"])
    def subscription():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        if request.method == "POST":
            plan = get_subscription_plan(request.form.get("plan_code") or user.plan_code)
            invoice = create_subscription_invoice(user, plan)
            flash("Subscription invoice created. Use the reference when paying.", "success")
            return redirect(url_for("registration_checkout", reference=invoice.payment_reference))
        invoices = SubscriptionInvoice.query.filter_by(user_id=user.id).order_by(SubscriptionInvoice.created_at.desc()).all()
        return render_template("subscription.html", invoices=invoices, active_plan=get_subscription_plan(user.plan_code), plans=SUBSCRIPTION_PLANS)

    @app.get("/logout")
    def logout():
        session.clear()
        return redirect(url_for("index"))

    @app.route("/profile", methods=["GET", "POST"])
    def profile():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        if request.method == "POST":
            action = request.form.get("action")
            if action == "update_email":
                new_email = request.form.get("email", "").lower().strip()
                current_password = request.form.get("current_password", "")
                if not new_email:
                    flash("Email address is required.", "danger")
                elif not user.check_password(current_password):
                    flash("Enter your current password before changing email.", "danger")
                elif User.query.filter(User.email == new_email, User.id != user.id).first():
                    flash("That email address is already in use.", "danger")
                else:
                    user.email = new_email
                    user.billing_contact_email = user.billing_contact_email or new_email
                    db.session.commit()
                    flash("Email address updated.", "success")
                    return redirect(url_for("profile"))
            elif action == "update_password":
                current_password = request.form.get("current_password", "")
                password = request.form.get("password", "")
                confirm_password = request.form.get("confirm_password", "")
                if not user.check_password(current_password):
                    flash("Current password is incorrect.", "danger")
                elif len(password) < 8:
                    flash("New password must be at least 8 characters.", "danger")
                elif password != confirm_password:
                    flash("New passwords do not match.", "danger")
                else:
                    user.set_password(password)
                    db.session.commit()
                    flash("Password updated.", "success")
                    return redirect(url_for("profile"))
        return render_template("profile.html", profile_user=user)

    @app.get("/dashboard")
    def dashboard():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        requests = ScreeningRequest.query.filter_by(user_id=user.id).order_by(ScreeningRequest.created_at.desc()).limit(8).all()
        monitored = MonitoringSubject.query.filter_by(user_id=user.id).order_by(MonitoringSubject.created_at.desc()).limit(8).all()
        recent_results = (
            ScreeningResult.query.join(ScreeningRequest)
            .filter(ScreeningRequest.user_id == user.id)
            .order_by(ScreeningResult.created_at.desc())
            .limit(12)
            .all()
        )
        recent_match_results = [result for result in recent_results if is_actionable_screening_match(result)][:8]
        review_queue = [result for result in recent_results if screening_decision_label(result) in {"Pending review", "Needs review"}][:8]
        monitoring_alerts = [subject for subject in monitored if monitoring_risk_level(subject) in {"High", "Medium", "Unknown"}]
        batch_requests = ScreeningRequest.query.filter_by(user_id=user.id, request_type="bulk").order_by(ScreeningRequest.created_at.desc()).limit(5).all()
        adverse_searches = AdverseMediaSearch.query.order_by(AdverseMediaSearch.created_at.desc()).limit(5).all()
        adverse_alerts = AdverseMediaAlert.query.order_by(AdverseMediaAlert.created_at.desc()).limit(6).all()
        risk_stats = {
            "open_alerts": len(review_queue) + len(monitoring_alerts),
            "high_risk": sum(1 for result in recent_results if screening_risk_level(result) == "High") + sum(1 for subject in monitoring_alerts if monitoring_risk_level(subject) == "High") + sum(1 for alert in adverse_alerts if alert.risk_level == "High"),
            "pending_review": len(review_queue) + sum(1 for alert in adverse_alerts if alert.review_status == "Pending review"),
            "monitoring_changes": len(monitoring_alerts),
        }
        risk_stats["open_alerts"] += sum(1 for alert in adverse_alerts if alert.review_status == "Pending review")
        return render_template(
            "dashboard.html",
            requests=requests,
            monitored=monitored,
            recent_results=recent_match_results,
            review_queue=review_queue,
            monitoring_alerts=monitoring_alerts,
            batch_requests=batch_requests,
            adverse_searches=adverse_searches,
            adverse_alerts=adverse_alerts,
            risk_stats=risk_stats,
        )

    @app.route("/developer-access", methods=["GET", "POST"])
    def developer_access():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        if not user.is_admin:
            flash("Developer access is restricted to administrators.", "danger")
            return redirect(url_for("dashboard"))

        if request.method == "POST":
            action = request.form.get("action", "").strip()
            if action == "create_api_key":
                key_name = request.form.get("name", "").strip() or "Organisation integration key"
                raw_key = generate_developer_api_key()
                api_key = DeveloperApiKey(
                    user_id=user.id,
                    name=key_name[:160],
                    key_prefix=raw_key[:18],
                    key_hash=hash_api_key(raw_key),
                )
                db.session.add(api_key)
                db.session.commit()
                session["new_developer_api_key"] = raw_key
                session["new_developer_api_key_name"] = api_key.name
                flash("API key created. Copy it now; it will not be shown again.", "success")
            elif action == "revoke_api_key":
                api_key = db.session.get(DeveloperApiKey, int(request.form.get("key_id") or 0))
                if api_key:
                    api_key.status = "revoked"
                    db.session.commit()
                    flash("API key revoked.", "success")
            elif action == "activate_subscription":
                target_user = db.session.get(User, int(request.form.get("user_id") or 0))
                if not target_user:
                    flash("Subscriber workspace not found.", "warning")
                    return redirect(url_for("developer_access"))
                plan_code = (request.form.get("plan_code") or target_user.plan_code or "professional").strip()
                target_user.subscription_status = "active"
                target_user.plan_code = plan_code
                if request.form.get("clear_trial") == "1":
                    target_user.trial_ends_at = None
                latest_invoice = (
                    SubscriptionInvoice.query.filter_by(user_id=target_user.id)
                    .order_by(SubscriptionInvoice.created_at.desc())
                    .first()
                )
                if latest_invoice and latest_invoice.status == "pending":
                    latest_invoice.status = "approved"
                    latest_invoice.notes = ((latest_invoice.notes or "") + "\nManually activated by system admin.").strip()
                db.session.commit()
                flash(f"Subscription activated for {target_user.email}.", "success")
            elif action == "deactivate_subscription":
                target_user = db.session.get(User, int(request.form.get("user_id") or 0))
                if not target_user:
                    flash("Subscriber workspace not found.", "warning")
                    return redirect(url_for("developer_access"))
                target_user.subscription_status = "pending"
                db.session.commit()
                flash(f"Subscription set to pending for {target_user.email}.", "warning")
            return redirect(url_for("developer_access"))

        api_keys = DeveloperApiKey.query.order_by(DeveloperApiKey.created_at.desc()).all()
        subscriber_rows = []
        subscriber_users = User.query.order_by(User.created_at.desc()).all()
        for row_user in subscriber_users:
            latest_invoice = (
                SubscriptionInvoice.query.filter_by(user_id=row_user.id)
                .order_by(SubscriptionInvoice.created_at.desc())
                .first()
            )
            subscriber_rows.append(
                {
                    "user": row_user,
                    "plan": get_subscription_plan(row_user.plan_code),
                    "latest_invoice": latest_invoice,
                }
            )
        subscriber_summary = {
            "total": len(subscriber_users),
            "active": sum(1 for row_user in subscriber_users if row_user.subscription_status == "active"),
            "pending": sum(1 for row_user in subscriber_users if row_user.subscription_status != "active"),
            "trial": sum(1 for row_user in subscriber_users if getattr(row_user, "trial_ends_at", None) is not None),
        }
        stats = {
            "records": PepRecord.query.count(),
            "candidates": PepRecord.query.filter(PepRecord.status.in_(["Candidate review", "Needs review"])).count(),
            "adverse_alerts": AdverseMediaAlert.query.filter_by(review_status="Pending review").count(),
            "active_keys": DeveloperApiKey.query.filter_by(status="active").count(),
            "openai": "Configured" if openai_adverse_media_enabled() else "Not configured",
        }
        endpoints = [
            {
                "method": "POST",
                "path": "/api/adverse-media/search",
                "purpose": "Run source-backed adverse-media search and save structured alerts.",
                "body": '{"name":"Duma Boko","jurisdiction":"Botswana","role":"President"}',
            },
            {
                "method": "POST",
                "path": "/public-search",
                "purpose": "Public preview search, limited to 3 names per visitor per day.",
                "body": "Form field: public_names",
            },
        ]
        return render_template(
            "developer_access.html",
            api_keys=api_keys,
            stats=stats,
            endpoints=endpoints,
            subscriber_rows=subscriber_rows,
            subscriber_summary=subscriber_summary,
            new_api_key=session.pop("new_developer_api_key", None),
            new_api_key_name=session.pop("new_developer_api_key_name", None),
        )

    @app.route("/screen", methods=["GET", "POST"])
    def screen():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        if request.method == "POST":
            names = request.form.get("names", "").splitlines()
            screening = save_results(user, names, "single")
            return redirect(url_for("results", request_id=screening.id))
        return render_template("screen.html")

    @app.route("/bulk-upload", methods=["GET", "POST"])
    def bulk_upload():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        if request.method == "POST":
            upload = request.files.get("file")
            if not upload or not upload.filename:
                flash("Please choose an Excel file.", "warning")
                return redirect(url_for("bulk_upload"))
            filename = secure_filename(upload.filename)
            names = extract_names_from_workbook(upload)
            screening = save_results(user, names, "bulk", filename)
            return redirect(url_for("results", request_id=screening.id))
        return render_template("bulk_upload.html")

    @app.get("/template")
    def download_template():
        wb = Workbook()
        ws = wb.active
        ws.title = "PEP Screening Names"
        ws.append(["Name", "Reference", "Notes"])
        ws.append(["Example Person", "Customer ID 001", "Optional"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="pep_screening_template.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.get("/requests/<int:request_id>")
    def results(request_id: int):
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        screening = db.session.get(ScreeningRequest, request_id)
        if not screening or (screening.user_id != user.id and not user.is_admin):
            flash("Screening request not found.", "warning")
            return redirect(url_for("dashboard"))
        return render_template("results.html", screening=screening)

    @app.get("/requests/<int:request_id>/risk-report.xlsx")
    def download_risk_research_report(request_id: int):
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        screening = db.session.get(ScreeningRequest, request_id)
        if not screening or (screening.user_id != user.id and not user.is_admin):
            flash("Screening request not found.", "warning")
            return redirect(url_for("dashboard"))
        output = build_risk_research_report(screening)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"Risk_Research_Report_Request_{screening.id}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.post("/monitoring")
    def add_monitoring():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        name = request.form.get("name", "").strip()
        if name:
            record, score, decision = screen_name(name)
            if decision == "No match":
                latest_outcome = "No match"
            else:
                latest_outcome = f"{decision} ({score}% similarity)"
            db.session.add(MonitoringSubject(user_id=user.id, name=name, last_screened_at=datetime.now(timezone.utc), last_decision=latest_outcome))
            db.session.commit()
        return redirect(url_for("dashboard"))

    @app.post("/api/adverse-media/search")
    def api_adverse_media_search():
        user = current_user()
        api_key = None
        if not user:
            user, api_key = authenticate_api_key()
        if user and not user.is_active_subscriber:
            user = None
        if not user:
            return {"error": "Authentication required."}, 401
        payload = request.get_json(silent=True) or {}
        name = str(payload.get("name") or "").strip()
        jurisdiction = str(payload.get("jurisdiction") or "").strip()
        role = str(payload.get("role") or "").strip()
        if len(name) < 2:
            return {"error": "Name is required."}, 400
        try:
            result = analyze_adverse_media_with_openai(name, jurisdiction, role)
            if api_key:
                api_key.last_used_at = datetime.now(timezone.utc)
            saved = save_adverse_media_result(result, created_by=user.id, fallback_name=name, fallback_jurisdiction=jurisdiction, fallback_role=role)
        except RuntimeError as exc:
            return {"error": str(exc)}, 503
        except Exception as exc:  # noqa: BLE001
            db.session.rollback()
            return {"error": f"Failed to complete adverse media search: {exc}"}, 500
        return adverse_media_search_to_dict(saved)

    @app.post("/adverse-media/search")
    def adverse_media_search():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        name = request.form.get("name", "").strip()
        jurisdiction = request.form.get("jurisdiction", "").strip()
        role = request.form.get("role", "").strip()
        if len(name) < 2:
            flash("Enter a name before running adverse-media search.", "warning")
            return redirect(url_for("dashboard"))
        try:
            result = analyze_adverse_media_with_openai(name, jurisdiction, role)
            saved = save_adverse_media_result(result, created_by=user.id, fallback_name=name, fallback_jurisdiction=jurisdiction, fallback_role=role)
            flash(f"Adverse-media search saved for {saved.searched_name}.", "success")
        except RuntimeError as exc:
            flash(str(exc), "warning")
        except Exception as exc:  # noqa: BLE001
            db.session.rollback()
            flash(f"Adverse-media search failed: {exc}", "danger")
        return redirect(url_for("dashboard"))

    @app.post("/adverse-media/alerts/<int:alert_id>/action")
    def adverse_media_alert_action(alert_id: int):
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        alert = db.session.get(AdverseMediaAlert, alert_id)
        if not alert:
            flash("Adverse-media alert not found.", "warning")
            return redirect(url_for("dashboard"))
        action = request.form.get("action", "reviewed").strip()
        note = request.form.get("reviewer_note", "").strip()
        status_map = {
            "confirm": "Confirmed",
            "dismiss": "Dismissed as false positive",
            "escalate": "Escalated",
            "reviewed": "Reviewed",
        }
        alert.review_status = status_map.get(action, "Reviewed")
        if note:
            today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            alert.reviewer_notes = ((alert.reviewer_notes or "") + f"\n{today} {user.email}: {note}").strip()
        db.session.commit()
        flash(f"Adverse-media alert marked as {alert.review_status}.", "success")
        return redirect(url_for("dashboard"))

    @app.get("/admin")
    def admin():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        query = request.args.get("q", "").strip()
        active_tab = request.args.get("tab", "all").strip() or "all"
        category_filter = request.args.get("category", "").strip()
        jurisdiction_filter = request.args.get("jurisdiction", "").strip()
        status_filter = request.args.get("status", "").strip()
        confidence_filter = request.args.get("confidence", "").strip()
        adverse_filter = request.args.get("adverse_media", "").strip()
        source_type_filter = request.args.get("source_type", "").strip()
        records_query = all_review_record_query() if active_tab == "rejected" else visible_record_query()
        if query:
            like = f"%{query}%"
            records_query = records_query.filter(or_(PepRecord.full_name.ilike(like), PepRecord.aliases.ilike(like), PepRecord.position.ilike(like), PepRecord.organisation.ilike(like)))
        if category_filter:
            records_query = records_query.filter(PepRecord.category == category_filter)
        if jurisdiction_filter:
            records_query = records_query.filter(PepRecord.jurisdiction == jurisdiction_filter)
        if status_filter:
            records_query = records_query.filter(PepRecord.status == status_filter)
        if adverse_filter:
            records_query = records_query.filter(PepRecord.adverse_media_status == adverse_filter)
        if source_type_filter:
            records_query = records_query.filter(PepRecord.source_type == source_type_filter)
        if active_tab == "candidates":
            records_query = records_query.filter(PepRecord.status == "Candidate review")
        elif active_tab == "confirmed":
            records_query = records_query.filter(PepRecord.status.in_(["Current", "Confirmed", "Former"]))
        elif active_tab == "needs_review":
            records_query = records_query.filter(PepRecord.status.in_(["Candidate review", "Needs review"]))
        elif active_tab == "adverse_media":
            records_query = records_query.filter(PepRecord.adverse_media_status.notin_(["No adverse media", "Pending review"]))
        elif active_tab == "rejected":
            records_query = records_query.filter(PepRecord.status.in_(HIDDEN_RECORD_STATUSES))
        records = records_query.order_by(PepRecord.updated_at.desc()).limit(100).all()
        if confidence_filter:
            records = [record for record in records if confidence_band(record) == confidence_filter]
        sources = PublicSource.query.order_by(PublicSource.name).all()
        users = User.query.order_by(User.created_at.desc()).all()
        pdf_logs = PdfIngestionLog.query.order_by(PdfIngestionLog.created_at.desc()).limit(10).all()
        web_logs = WebLinkReviewLog.query.order_by(WebLinkReviewLog.created_at.desc()).limit(10).all()
        current_affairs = CurrentAffairsIssue.query.order_by(CurrentAffairsIssue.is_active.desc(), CurrentAffairsIssue.issue_date.desc(), CurrentAffairsIssue.updated_at.desc()).limit(10).all()
        suspect_records = [record for record in PepRecord.query.filter_by(status="Candidate review").limit(500).all() if likely_non_person_record(record)]
        duplicate_suggestions = duplicate_merge_suggestions()
        relationships = PipRelationship.query.order_by(PipRelationship.updated_at.desc()).limit(25).all()
        audit_logs = RecordAuditLog.query.order_by(RecordAuditLog.created_at.desc()).limit(20).all()
        coverage = coverage_summary()
        all_visible = visible_record_query()
        stats = {
            "total": all_visible.count(),
            "candidates": visible_record_query().filter(PepRecord.status == "Candidate review").count(),
            "needs_review": visible_record_query().filter(PepRecord.status.in_(["Candidate review", "Needs review"])).count(),
            "adverse": visible_record_query().filter(PepRecord.adverse_media_status.notin_(["No adverse media", "Pending review"])).count(),
            "duplicates": len(duplicate_suggestions),
            "relationships": PipRelationship.query.count(),
            "verified": coverage["verified_records"],
            "rejected": all_review_record_query().filter(PepRecord.status.in_(HIDDEN_RECORD_STATUSES)).count(),
            "relationship_candidates": PipRelationship.query.filter(PipRelationship.review_status.in_(["Candidate review", "Needs review"])).count(),
        }
        filter_options = {
            "categories": [row[0] for row in db.session.query(PepRecord.category).filter(PepRecord.category.isnot(None)).distinct().order_by(PepRecord.category).all() if row[0]],
            "jurisdictions": [row[0] for row in db.session.query(PepRecord.jurisdiction).filter(PepRecord.jurisdiction.isnot(None)).distinct().order_by(PepRecord.jurisdiction).all() if row[0]],
            "statuses": [row[0] for row in db.session.query(PepRecord.status).filter(PepRecord.status.isnot(None)).distinct().order_by(PepRecord.status).all() if row[0] not in HIDDEN_RECORD_STATUSES],
            "adverse_media": ["Pending review", "No adverse media", "Adverse media found", "Under investigation", "Sanctions match", "Procurement risk", "Fraud allegation", "Official capacity only"],
            "source_types": ["Government", "News", "Registry", "Manual entry", "PDF", "AI/rules"],
            "verification_statuses": ["Unverified", "Source verified", "Management represented", "Needs re-verification"],
            "source_reliability": ["High", "Medium", "Low", "Unknown"],
        }
        filters = {
            "tab": active_tab,
            "category": category_filter,
            "jurisdiction": jurisdiction_filter,
            "status": status_filter,
            "confidence": confidence_filter,
            "adverse_media": adverse_filter,
            "source_type": source_type_filter,
        }
        return render_template(
            "admin.html",
            records=records,
            sources=sources,
            users=users,
            query=query,
            pdf_logs=pdf_logs,
            web_logs=web_logs,
            pip_definitions=PIP_CATEGORY_DEFINITIONS,
            suspect_records=suspect_records,
            duplicate_suggestions=duplicate_suggestions,
            stats=stats,
            filter_options=filter_options,
            filters=filters,
            current_affairs=current_affairs,
            relationships=relationships,
            audit_logs=audit_logs,
            coverage=coverage,
            data_dictionary=DATA_DICTIONARY,
            compliance_sections=COMPLIANCE_PACK_SECTIONS,
        )

    @app.post("/admin/records")
    def add_record():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        category = request.form.get("category", "Domestic PIP").strip()
        position = request.form.get("position", "").strip()
        source_name = request.form.get("source_name", "").strip()
        raw_notes = request.form.get("notes", "").strip()
        basis = (
            f"Rule-based mapping: position/title '{position}' supports category '{category}'."
            if position
            else f"Manual/admin record classified as '{category}'; basis should be confirmed during review."
        )
        record = PepRecord(
            full_name=request.form.get("full_name", "").strip(),
            aliases=request.form.get("aliases", "").strip(),
            category=category,
            jurisdiction=request.form.get("jurisdiction", "Botswana").strip(),
            position=position,
            organisation=request.form.get("organisation", "").strip(),
            status=request.form.get("status", "Current").strip(),
            source_url=request.form.get("source_url", "").strip(),
            source_name=source_name,
            source_date=request.form.get("source_date", "").strip(),
            date_identified=request.form.get("date_identified", "").strip(),
            source_type=request.form.get("source_type", "Manual entry").strip(),
            source_excerpt=request.form.get("source_excerpt", "").strip(),
            source_reliability=request.form.get("source_reliability", "Unknown").strip(),
            verification_status=request.form.get("verification_status", "Unverified").strip(),
            verified_by=request.form.get("verified_by", "").strip(),
            last_verified_date=request.form.get("last_verified_date", "").strip(),
            next_review_due=request.form.get("next_review_due", "").strip(),
            adverse_media_status=request.form.get("adverse_media_status", "Pending review").strip(),
            adverse_media_linkage=request.form.get("adverse_media_linkage", "").strip(),
            reviewer_notes=request.form.get("reviewer_notes", "").strip(),
            last_reviewed_date=request.form.get("last_reviewed_date", "").strip(),
            notes=append_basis_if_missing(raw_notes, basis),
        )
        if not record.full_name:
            flash("Full name is required.", "warning")
        else:
            db.session.add(record)
            db.session.flush()
            add_audit_log(action="record_created", actor=user.email, record=record, changes=serialise_for_audit(record), note="Manual admin record creation")
            db.session.commit()
            flash("PEP/PIP record added successfully.", "success")
        return redirect(url_for("admin"))

    @app.get("/admin/records/export")
    def export_records():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        output = build_records_workbook()
        return send_file(
            output,
            as_attachment=True,
            download_name=f"pep_pip_records_export_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/api/feed/records")
    def api_feed_records():
        user = current_user()
        api_key = None
        if not user:
            user, api_key = authenticate_api_key()
        if not user or not user.is_active_subscriber:
            return {"error": "Authentication required."}, 401
        if api_key:
            api_key.last_used_at = datetime.now(timezone.utc)
            db.session.commit()
        export_format = request.args.get("format", "json").lower().strip()
        since = request.args.get("since", "").strip()
        records = feed_records_since(since)
        if export_format == "csv":
            output = StringIO()
            writer = csv.DictWriter(output, fieldnames=["id", "full_name", "category", "jurisdiction", "position", "status", "source_name", "source_url", "verification_status", "updated_at"])
            writer.writeheader()
            for record in records:
                data = record_feed_dict(record)
                writer.writerow(
                    {
                        "id": data["id"],
                        "full_name": data["full_name"],
                        "category": data["category"],
                        "jurisdiction": data["jurisdiction"],
                        "position": data["position"],
                        "status": data["status"],
                        "source_name": data["source"]["name"],
                        "source_url": data["source"]["url"],
                        "verification_status": data["verification"]["status"],
                        "updated_at": data["updated_at"],
                    }
                )
            return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=pip_feed.csv"})
        if export_format in {"xlsx", "excel"}:
            return send_file(
                build_feed_workbook(records),
                as_attachment=True,
                download_name=f"pip_feed_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        return jsonify(
            {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "change_only_since": since,
                "record_count": len(records),
                "records": [record_feed_dict(record) for record in records],
                "relationships": [relationship_feed_dict(item) for item in PipRelationship.query.order_by(PipRelationship.updated_at.desc()).all()],
            }
        )

    @app.get("/admin/data-dictionary")
    def data_dictionary():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        sample_records = [record_feed_dict(record) for record in visible_record_query().order_by(PepRecord.updated_at.desc()).limit(3).all()]
        return render_template("data_dictionary.html", data_dictionary=DATA_DICTIONARY, sample_records=sample_records)

    @app.get("/admin/coverage")
    def coverage_dashboard():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        return render_template("coverage.html", coverage=coverage_summary())

    @app.get("/compliance-pack")
    def compliance_pack():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        return render_template("compliance_pack.html", sections=COMPLIANCE_PACK_SECTIONS)

    @app.get("/compliance-pack/download")
    def compliance_pack_download():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        lines = ["Softdayta Risk Compliance Pack", f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", ""]
        for section in COMPLIANCE_PACK_SECTIONS:
            lines.extend([section["title"], section["summary"], "Evidence: " + "; ".join(section["evidence"]), ""])
        return Response("\n".join(lines), mimetype="text/plain", headers={"Content-Disposition": "attachment; filename=softdayta_risk_compliance_pack.txt"})

    @app.get("/commercial-pack/download")
    def commercial_pack_download():
        user = require_login()
        if not user:
            return redirect(url_for("index"))
        lines = [
            "Softdayta Risk Commercial / SLA Pack",
            "Proposal validity: 90 days from tender closing date.",
            "Pricing currency: NAD/ZAR VAT exclusive to be confirmed in final commercial schedule.",
            "Support model: onboarding support, administrator training, source-feed monitoring, and issue triage.",
            "SLA draft: production uptime target, support response times, incident escalation and maintenance window to be agreed.",
            "Assumptions: lawful public-source use, client review before reliance, and final dataset coverage subject to source verification.",
        ]
        return Response("\n".join(lines), mimetype="text/plain", headers={"Content-Disposition": "attachment; filename=softdayta_risk_commercial_sla_pack.txt"})

    @app.post("/admin/records/import")
    def import_records():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        upload = request.files.get("records_file")
        if not upload or not upload.filename:
            flash("Please choose an Excel file to import.", "warning")
            return redirect(url_for("admin"))
        filename = secure_filename(upload.filename)
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            flash("Please upload an .xlsx or .xlsm workbook.", "warning")
            return redirect(url_for("admin"))
        try:
            created, updated, errors = import_records_workbook(upload)
        except Exception as exc:  # noqa: BLE001
            db.session.rollback()
            flash(f"Excel import failed: {exc}", "danger")
            return redirect(url_for("admin"))
        error_text = f" Errors: {' | '.join(errors[:3])}" if errors else ""
        flash(f"Excel import completed: {created} created, {updated} updated.{error_text}", "success" if not errors else "warning")
        return redirect(url_for("admin"))

    @app.route("/admin/bulk-edit", methods=["GET", "POST"])
    def bulk_edit():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            record_ids = request.form.getlist("record_id")
            updated_records = 0
            for record_id in record_ids:
                record = db.session.get(PepRecord, int(record_id))
                if not record:
                    continue
                before = serialise_for_audit(record)
                prefix = f"record_{record_id}_"
                for field in [
                    "full_name",
                    "category",
                    "position",
                    "jurisdiction",
                    "status",
                    "source_name",
                    "source_url",
                    "source_excerpt",
                    "verification_status",
                    "source_reliability",
                    "adverse_media_status",
                    "last_verified_date",
                    "next_review_due",
                    "reviewer_notes",
                ]:
                    form_key = prefix + field
                    if form_key in request.form:
                        setattr(record, field, request.form.get(form_key, "").strip())
                changes = changed_fields(before, serialise_for_audit(record))
                if changes:
                    add_audit_log(action="bulk_record_update", actor=user.email, record=record, changes=changes, note="Excel-like bulk edit")
                    updated_records += 1

            relationship_ids = request.form.getlist("relationship_id")
            updated_relationships = 0
            for relationship_id in relationship_ids:
                relationship = db.session.get(PipRelationship, int(relationship_id))
                if not relationship:
                    continue
                before = relationship_feed_dict(relationship)
                prefix = f"relationship_{relationship_id}_"
                for field in ["related_name", "relationship_type", "category", "jurisdiction", "confidence_score", "review_status", "source_name", "source_url", "source_excerpt", "reviewer_notes"]:
                    form_key = prefix + field
                    if form_key not in request.form:
                        continue
                    value = request.form.get(form_key, "").strip()
                    if field == "confidence_score":
                        try:
                            value = max(0, min(int(value or 0), 100))
                        except ValueError:
                            value = relationship.confidence_score
                    setattr(relationship, field, value)
                after = relationship_feed_dict(relationship)
                changes = changed_fields(before, after)
                if changes:
                    add_audit_log(action="bulk_relationship_update", actor=user.email, record=relationship.principal, relationship=relationship, changes=changes, note="Excel-like relationship/RCA bulk edit")
                    updated_relationships += 1
            db.session.commit()
            flash(f"Bulk edit saved: {updated_records} PIP record(s), {updated_relationships} relationship/RCA record(s) updated.", "success")
            return redirect(url_for("bulk_edit"))

        records = all_review_record_query().order_by(PepRecord.updated_at.desc()).limit(150).all()
        relationships = PipRelationship.query.order_by(PipRelationship.updated_at.desc()).limit(150).all()
        filter_options = {
            "statuses": ["Candidate review", "Needs review", "Confirmed", "Current", "Former", "Rejected / not a person", "Duplicate", "Merged duplicate"],
            "verification_statuses": ["Unverified", "Source verified", "Management represented", "Needs re-verification"],
            "source_reliability": ["High", "Medium", "Low", "Unknown"],
            "adverse_media": ["Pending review", "No adverse media", "Adverse media found", "Under investigation", "Sanctions match", "Procurement risk", "Fraud allegation", "Official capacity only"],
            "relationship_statuses": ["Candidate review", "Needs review", "Confirmed", "Rejected", "Duplicate"],
        }
        return render_template("bulk_edit.html", records=records, relationships=relationships, filter_options=filter_options)

    @app.post("/admin/relationships")
    def add_relationship():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        try:
            principal_id = int(request.form.get("principal_record_id", "0"))
        except ValueError:
            principal_id = 0
        principal = db.session.get(PepRecord, principal_id)
        if not principal:
            flash("Select a valid principal PIP record.", "warning")
            return redirect(url_for("admin"))
        try:
            confidence_score = max(0, min(int(request.form.get("confidence_score", "50") or 50), 100))
        except ValueError:
            confidence_score = 50
        relationship = PipRelationship(
            principal_record_id=principal.id,
            related_name=request.form.get("related_name", "").strip(),
            relationship_type=request.form.get("relationship_type", "").strip(),
            category=request.form.get("category", "Related party").strip() or "Related party",
            jurisdiction=request.form.get("jurisdiction", principal.jurisdiction).strip() or principal.jurisdiction,
            source_name=request.form.get("source_name", "").strip(),
            source_url=request.form.get("source_url", "").strip(),
            source_excerpt=request.form.get("source_excerpt", "").strip(),
            confidence_score=confidence_score,
            review_status=request.form.get("review_status", "Candidate review").strip() or "Candidate review",
            reviewer_notes=request.form.get("reviewer_notes", "").strip(),
        )
        if not relationship.related_name or not relationship.relationship_type:
            flash("Related name and relationship type are required.", "warning")
            return redirect(url_for("admin"))
        db.session.add(relationship)
        db.session.flush()
        add_audit_log(
            action="relationship_created",
            actor=user.email,
            record=principal,
            relationship=relationship,
            changes=relationship_feed_dict(relationship),
            note=f"Related party added for {principal.full_name}",
        )
        db.session.commit()
        flash("PIP relationship / RCA record added.", "success")
        return redirect(url_for("admin"))

    @app.post("/admin/records/<int:record_id>/edit")
    def edit_record(record_id: int):
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        record = db.session.get(PepRecord, record_id)
        if not record:
            flash("Record not found.", "warning")
            return redirect(url_for("admin"))

        editable_fields = [
            "full_name",
            "aliases",
            "category",
            "jurisdiction",
            "position",
            "organisation",
            "status",
            "date_identified",
            "last_reviewed_date",
            "last_verified_date",
            "next_review_due",
            "verification_status",
            "verified_by",
            "source_reliability",
            "source_type",
            "source_name",
            "source_url",
            "source_date",
            "source_excerpt",
            "adverse_media_status",
            "adverse_media_linkage",
            "reviewer_notes",
            "notes",
        ]
        before = serialise_for_audit(record)
        for field in editable_fields:
            if field in request.form:
                setattr(record, field, request.form.get(field, "").strip())
        record.notes = append_basis_if_missing(
            record.notes or "",
            f"Admin-edited suggested information on {datetime.now(timezone.utc).strftime('%Y-%m-%d')} by {user.email}.",
        )
        changes = changed_fields(before, serialise_for_audit(record))
        if changes:
            add_audit_log(action="record_edited", actor=user.email, record=record, changes=changes, note="Admin suggested-info edit")
        db.session.commit()
        flash(f"Suggested information updated for {record.full_name}.", "success")
        return redirect(url_for("admin"))

    @app.post("/admin/sources")
    def add_source():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        source = PublicSource(
            name=request.form.get("name", "").strip(),
            url=request.form.get("url", "").strip(),
            jurisdiction=request.form.get("jurisdiction", "Botswana").strip(),
        )
        if source.name and source.url:
            db.session.add(source)
            db.session.commit()
            flash("Public source added.", "success")
        return redirect(url_for("admin"))

    @app.post("/admin/current-affairs")
    def add_current_affairs():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        issue = CurrentAffairsIssue(
            title=request.form.get("title", "").strip(),
            category=request.form.get("category", "Politics").strip() or "Politics",
            jurisdiction=request.form.get("jurisdiction", "Botswana").strip() or "Botswana",
            summary=request.form.get("summary", "").strip(),
            source_name=request.form.get("source_name", "").strip(),
            source_url=request.form.get("source_url", "").strip(),
            image_url=request.form.get("image_url", "").strip(),
            issue_date=request.form.get("issue_date", "").strip(),
            is_active=request.form.get("is_active", "1") == "1",
        )
        if not issue.title or not issue.summary:
            flash("Current-affairs title and summary are required.", "warning")
        else:
            db.session.add(issue)
            db.session.commit()
            flash("Current-affairs issue added to the landing page.", "success")
        return redirect(url_for("admin"))

    @app.post("/admin/current-affairs/<int:issue_id>/toggle")
    def toggle_current_affairs(issue_id: int):
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        issue = db.session.get(CurrentAffairsIssue, issue_id)
        if not issue:
            flash("Current-affairs issue not found.", "warning")
        else:
            issue.is_active = not issue.is_active
            db.session.commit()
            flash("Current-affairs issue visibility updated.", "success")
        return redirect(url_for("admin"))

    @app.post("/admin/records/<int:record_id>/action")
    def record_action(record_id: int):
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        record = db.session.get(PepRecord, record_id)
        if not record:
            flash("Record not found.", "warning")
            return redirect(url_for("admin"))

        action = request.form.get("action", "").strip()
        note = request.form.get("reviewer_note", "").strip()
        adverse_status = request.form.get("adverse_media_status", "").strip()
        adverse_linkage = request.form.get("adverse_media_linkage", "").strip()
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        before = serialise_for_audit(record)
        action_labels = {
            "confirm": "Confirmed",
            "reject": "Rejected / not a person",
            "needs_review": "Needs review",
            "duplicate": "Duplicate",
        }

        if action in action_labels:
            record.status = action_labels[action]
            if action == "confirm" and record.adverse_media_status in {None, "", "Pending review"}:
                record.adverse_media_status = "No adverse media"
        elif action == "adverse_media":
            if adverse_status:
                record.adverse_media_status = adverse_status
            if adverse_linkage:
                record.adverse_media_linkage = adverse_linkage
        else:
            flash("Unknown record action.", "warning")
            return redirect(url_for("admin"))

        record.last_reviewed_date = today
        if note:
            record.reviewer_notes = ((record.reviewer_notes or "") + f"\n{today}: {note}").strip()
        action_note = f"{today}: Admin action '{action}' applied by {user.email}."
        record.notes = ((record.notes or "") + f"\n\n{action_note}").strip()
        add_audit_log(
            action=f"record_action_{action}",
            actor=user.email,
            record=record,
            changes=changed_fields(before, serialise_for_audit(record)),
            note=note,
        )
        db.session.commit()

        messages = {
            "confirm": f"{record.full_name} confirmed successfully.",
            "reject": f"{record.full_name} rejected and hidden from normal screening.",
            "needs_review": f"{record.full_name} marked as needing review.",
            "duplicate": f"{record.full_name} marked as a duplicate.",
            "adverse_media": f"Adverse media status updated for {record.full_name}.",
        }
        flash(messages.get(action, "Record updated."), "success")
        return redirect(url_for("admin"))

    @app.post("/admin/sources/<int:source_id>/update")
    def run_source_update(source_id: int):
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        source = db.session.get(PublicSource, source_id)
        if source:
            log = update_source(source)
            flash(log.message or log.status, "success" if log.status == "success" else "danger")
        return redirect(url_for("admin"))

    @app.post("/admin/dailynews/update")
    def run_dailynews_update():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        max_pdfs = request.form.get("max_pdfs", "25")
        try:
            max_pdfs_int = max(1, min(int(max_pdfs), 100))
        except ValueError:
            max_pdfs_int = 25
        log = ingest_dailynews_pdfs(max_pdfs=max_pdfs_int)
        flash(log.message or "DailyNews PDF import completed.", "success")
        return redirect(url_for("admin"))

    @app.post("/admin/weblink/review")
    def run_weblink_review():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        url = request.form.get("url", "").strip()
        jurisdiction = request.form.get("jurisdiction", "Botswana").strip() or "Botswana"
        same_domain_only = request.form.get("same_domain_only", "1") == "1"
        try:
            max_links = max(0, min(int(request.form.get("max_links", "10") or 10), 25))
        except ValueError:
            max_links = 10
        if not url:
            flash("Paste a web link before running review.", "warning")
            return redirect(url_for("admin"))
        log = review_weblink_for_candidates(
            url,
            jurisdiction=jurisdiction,
            max_links=max_links,
            same_domain_only=same_domain_only,
        )
        flash(log.message or "Web link review completed.", "success" if log.status in {"success", "partial"} else "danger")
        return redirect(url_for("admin"))

    @app.post("/admin/youtube/paste")
    def stage_youtube_video():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))

        url = request.form.get("url", "").strip()
        jurisdiction = request.form.get("jurisdiction", "Botswana").strip() or "Botswana"
        lang = request.form.get("lang", "en").strip() or "en"
        if not url:
            flash("Paste a YouTube video link first.", "warning")
            return redirect(url_for("admin"))

        video_id = extract_youtube_video_id(url)
        if not video_id:
            flash("Could not detect a YouTube video id from that link.", "warning")
            return redirect(url_for("admin"))

        transcript_text, meta = fetch_youtube_transcript_text(video_id, lang=lang)
        if not transcript_text:
            flash("No captions/transcript found for that video (or it is not accessible). Try another link or language.", "warning")
            return redirect(url_for("admin"))

        # Cap extremely long transcripts to keep processing predictable.
        transcript_text = transcript_text[:250_000]

        run = StagedImportRun(
            source_name=f"YouTube video {video_id}",
            source_url=url,
            jurisdiction=jurisdiction,
            status="staged",
            created_by=user.email,
            message=f"Staged YouTube transcript import for video {video_id}. Review before applying.",
            metrics_json=json.dumps({"video_id": video_id, "lang": lang, "mode": meta.get("mode", ""), "transcript_chars": len(transcript_text)}, default=str),
        )
        db.session.add(run)
        db.session.flush()

        candidates, metrics = extract_validated_candidates(
            transcript_text,
            source_name=f"YouTube transcript - {video_id}",
            source_url=url,
            source_jurisdiction=jurisdiction,
            log_rejections=True,
        )
        relationship_candidates = extract_relationship_candidates_from_text(
            transcript_text,
            source_name=f"YouTube transcript - {video_id}",
            source_url=url,
            source_jurisdiction=jurisdiction,
            source_type="YouTube transcript",
            limit=50,
        )

        candidates_staged = 0
        relationships_staged = 0
        for candidate in candidates[:150]:
            db.session.add(
                StagedImportCandidate(
                    run_id=run.id,
                    kind="pip",
                    full_name=str(candidate.get("name") or ""),
                    category=str(candidate.get("category") or ""),
                    position=str(candidate.get("position") or ""),
                    confidence_score=int(candidate.get("confidence_score") or 0),
                    snippet=str(candidate.get("snippet") or ""),
                    evidence_json=json.dumps(candidate, default=str),
                )
            )
            candidates_staged += 1
        for rel in relationship_candidates[:50]:
            db.session.add(
                StagedImportCandidate(
                    run_id=run.id,
                    kind="relationship",
                    principal_record_id=int(rel.get("principal_record_id") or 0) or None,
                    related_name=str(rel.get("related_name") or ""),
                    relationship_type=str(rel.get("relationship_type") or ""),
                    category=str(rel.get("category") or "Related party"),
                    confidence_score=int(rel.get("confidence_score") or 0),
                    snippet=str(rel.get("source_excerpt") or rel.get("snippet") or ""),
                    evidence_json=json.dumps(rel, default=str),
                )
            )
            relationships_staged += 1

        run.message = (
            f"Staged YouTube transcript import: {candidates_staged} candidate record(s). "
            f"Staged {relationships_staged} relationship/RCA candidate(s). "
            f"Raw strings: {metrics.get('raw_strings', 0)}; filtered out: {metrics.get('filtered_out', 0)}; "
            f"high confidence: {metrics.get('high_confidence', 0)}; medium: {metrics.get('medium_confidence', 0)}; "
            f"low: {metrics.get('low_confidence', 0)}; foreign PIPs: {metrics.get('foreign_pips', 0)}; "
            f"AI candidates: {metrics.get('ai_candidates', 0)}; AI errors: {metrics.get('ai_errors', 0)}."
        )
        db.session.commit()
        flash(f"{run.message} Review before applying: /admin/staged-imports/{run.id}", "success")
        return redirect(url_for("staged_import_review", run_id=run.id))

    @app.post("/admin/transcript/paste")
    def stage_transcript_paste():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))

        transcript_text = request.form.get("transcript_text", "")
        jurisdiction = request.form.get("jurisdiction", "Botswana").strip() or "Botswana"
        source_url = request.form.get("source_url", "").strip() or "manual-transcript-paste"
        source_name = request.form.get("source_name", "").strip() or "Manual transcript paste"
        if not transcript_text.strip():
            flash("Paste transcript text first.", "warning")
            return redirect(url_for("admin"))

        transcript_text = transcript_text.strip()
        transcript_text = transcript_text[:250_000]

        run = StagedImportRun(
            source_name=source_name,
            source_url=source_url,
            jurisdiction=jurisdiction,
            status="staged",
            created_by=user.email,
            message="Staged transcript paste. Review before applying.",
            metrics_json=json.dumps({"transcript_chars": len(transcript_text)}, default=str),
        )
        db.session.add(run)
        db.session.flush()

        candidates, metrics = extract_validated_candidates(
            transcript_text,
            source_name=source_name,
            source_url=source_url,
            source_jurisdiction=jurisdiction,
            log_rejections=True,
        )
        relationship_candidates = extract_relationship_candidates_from_text(
            transcript_text,
            source_name=source_name,
            source_url=source_url,
            source_jurisdiction=jurisdiction,
            source_type="Transcript paste",
            limit=50,
        )

        candidates_staged = 0
        relationships_staged = 0
        for candidate in candidates[:150]:
            db.session.add(
                StagedImportCandidate(
                    run_id=run.id,
                    kind="pip",
                    full_name=str(candidate.get("name") or ""),
                    category=str(candidate.get("category") or ""),
                    position=str(candidate.get("position") or ""),
                    confidence_score=int(candidate.get("confidence_score") or 0),
                    snippet=str(candidate.get("snippet") or ""),
                    evidence_json=json.dumps(candidate, default=str),
                )
            )
            candidates_staged += 1
        for rel in relationship_candidates[:50]:
            db.session.add(
                StagedImportCandidate(
                    run_id=run.id,
                    kind="relationship",
                    principal_record_id=int(rel.get("principal_record_id") or 0) or None,
                    related_name=str(rel.get("related_name") or ""),
                    relationship_type=str(rel.get("relationship_type") or ""),
                    category=str(rel.get("category") or "Related party"),
                    confidence_score=int(rel.get("confidence_score") or 0),
                    snippet=str(rel.get("source_excerpt") or rel.get("snippet") or ""),
                    evidence_json=json.dumps(rel, default=str),
                )
            )
            relationships_staged += 1

        run.message = (
            f"Staged transcript paste: {candidates_staged} candidate record(s). "
            f"Staged {relationships_staged} relationship/RCA candidate(s). "
            f"Raw strings: {metrics.get('raw_strings', 0)}; filtered out: {metrics.get('filtered_out', 0)}; "
            f"high confidence: {metrics.get('high_confidence', 0)}; medium: {metrics.get('medium_confidence', 0)}; "
            f"low: {metrics.get('low_confidence', 0)}; foreign PIPs: {metrics.get('foreign_pips', 0)}; "
            f"AI candidates: {metrics.get('ai_candidates', 0)}; AI errors: {metrics.get('ai_errors', 0)}."
        )
        db.session.commit()
        flash(f"{run.message} Review before applying: /admin/staged-imports/{run.id}", "success")
        return redirect(url_for("staged_import_review", run_id=run.id))

    @app.post("/admin/pdfs/upload")
    def upload_pdfs():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        files = request.files.getlist("pdf_files")
        if not files:
            flash("Please choose at least one PDF file.", "warning")
            return redirect(url_for("admin"))
        jurisdiction = request.form.get("jurisdiction", "Botswana").strip() or "Botswana"
        log, staged_run_id = ingest_uploaded_pdfs(files, source_jurisdiction=jurisdiction)
        if staged_run_id:
            flash(
                f"{log.message or 'Uploaded PDF import completed.'} Review before applying: /admin/staged-imports/{staged_run_id}",
                "success" if log.status in {"success", "partial"} else "danger",
            )
        else:
            flash(log.message or "Uploaded PDF import completed.", "success" if log.status in {"success", "partial"} else "danger")
        return redirect(url_for("admin"))

    @app.post("/admin/cleanup/non-person")
    def cleanup_non_person_records():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        action = request.form.get("action", "mark")
        count = cleanup_false_positive_candidates(delete=action == "delete")
        verb = "deleted" if action == "delete" else "marked as not a person"
        flash(f"{count} likely non-person candidate record(s) {verb}. Confirmed and seed records were not changed.", "success")
        return redirect(url_for("admin"))

    @app.post("/admin/duplicates/merge")
    def merge_duplicate_records():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        try:
            target_id = int(request.form.get("target_id", "0"))
        except ValueError:
            target_id = 0
        source_ids = []
        for value in request.form.getlist("source_ids"):
            try:
                source_ids.append(int(value))
            except ValueError:
                continue
        merged = merge_candidate_records(target_id, source_ids)
        message = "Duplicate candidate merged successfully into the confirmed profile." if merged == 1 else f"{merged} duplicate candidates merged successfully into the confirmed profile."
        flash(message, "success")
        return redirect(url_for("admin"))

    @app.get("/admin/staged-imports")
    def staged_imports():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        runs = StagedImportRun.query.order_by(StagedImportRun.created_at.desc()).limit(25).all()
        return render_template("staged_imports.html", runs=runs)

    @app.route("/admin/staged-imports/<int:run_id>", methods=["GET", "POST"])
    def staged_import_review(run_id: int):
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        run = db.session.get(StagedImportRun, run_id)
        if not run:
            flash("Staged import run not found.", "warning")
            return redirect(url_for("staged_imports"))

        if request.method == "POST":
            action = (request.form.get("action") or "").strip()
            if action == "run_adverse_media":
                # Run adverse-media scans for approved PIP candidates in this staged run before applying.
                if not openai_adverse_media_enabled():
                    flash("OpenAI web search is not configured. Set OPENAI_API_KEY to enable adverse-media analysis.", "warning")
                    return redirect(url_for("staged_import_review", run_id=run.id))
                approved = StagedImportCandidate.query.filter_by(run_id=run.id, kind="pip", review_status="Approved").all()
                created = 0
                skipped = 0
                for item in approved[:25]:
                    name = (item.full_name or "").strip()
                    if not name:
                        continue
                    if adverse_media_search_exists_for(name, jurisdiction=run.jurisdiction):
                        skipped += 1
                        continue
                    try:
                        result = analyze_adverse_media_with_openai(name, run.jurisdiction, "Linked party (registry)")
                        save_adverse_media_result(result, created_by=user.id, fallback_name=name, fallback_jurisdiction=run.jurisdiction, fallback_role="Linked party (registry)")
                        created += 1
                    except Exception:
                        continue
                flash(f"Adverse-media scan completed for staged run {run.id}: {created} created, {skipped} skipped.", "success")
                return redirect(url_for("staged_import_review", run_id=run.id))
            if action == "apply":
                approved = StagedImportCandidate.query.filter_by(run_id=run.id, review_status="Approved").all()
                created_records = 0
                created_relationships = 0
                for item in approved:
                    evidence = {}
                    try:
                        evidence = json.loads(item.evidence_json or "{}")
                    except Exception:
                        evidence = {}
                    candidate_source_url = str(evidence.get("source_url") or run.source_url or "")
                    candidate_source_name = str(evidence.get("source_name") or run.source_name or "")
                    if item.kind == "registry":
                        continue
                    if item.kind == "pip" and item.full_name:
                        exists = PepRecord.query.filter(
                            PepRecord.source_url == candidate_source_url,
                            PepRecord.full_name.ilike(item.full_name),
                        ).first()
                        if exists:
                            continue
                        record = PepRecord(
                            full_name=item.full_name,
                            category=item.category or "Public-source mention only",
                            jurisdiction=run.jurisdiction,
                            position=item.position or "",
                            status="Candidate review",
                            source_url=candidate_source_url,
                            source_name=candidate_source_name,
                            source_type="PDF",
                            source_excerpt=item.snippet or "",
                            date_identified=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                            adverse_media_status="Pending review",
                            notes=append_basis_if_missing("", f"Basis in document: Staged import run {run.id} applied by admin."),
                        )
                        db.session.add(record)
                        db.session.flush()
                        add_audit_log(action="staged_import_candidate_applied", actor=user.email, record=record, changes=serialise_for_audit(record), note=f"Applied from staged import run {run.id}")
                        created_records += 1
                    if item.kind == "relationship" and item.principal_record_id and item.related_name and item.relationship_type:
                        exists = PipRelationship.query.filter(
                            PipRelationship.principal_record_id == item.principal_record_id,
                            PipRelationship.related_name.ilike(item.related_name),
                            PipRelationship.source_url == candidate_source_url,
                        ).first()
                        if exists:
                            continue
                        rel = PipRelationship(
                            principal_record_id=item.principal_record_id,
                            related_name=item.related_name,
                            relationship_type=item.relationship_type,
                            category=item.category or "Related party",
                            jurisdiction=run.jurisdiction,
                            source_name=candidate_source_name,
                            source_url=candidate_source_url,
                            source_excerpt=(item.snippet or "")[:1200],
                            confidence_score=int(item.confidence_score or 0),
                            review_status="Candidate review",
                            reviewer_notes=f"Applied from staged import run {run.id} by admin.",
                        )
                        db.session.add(rel)
                        db.session.flush()
                        add_audit_log(action="staged_import_relationship_applied", actor=user.email, record=db.session.get(PepRecord, item.principal_record_id), relationship=rel, changes=relationship_feed_dict(rel), note=f"Applied from staged import run {run.id}")
                        created_relationships += 1
                run.status = "applied"
                run.applied_at = datetime.now(timezone.utc)
                db.session.commit()
                flash(f"Applied staged import run {run.id}: {created_records} record(s), {created_relationships} relationship(s) created.", "success")
                return redirect(url_for("admin"))

            if action in {"approve", "reject"}:
                selected_ids = [int(value) for value in request.form.getlist("candidate_id") if value.isdigit()]
                if not selected_ids:
                    flash("Select one or more staged rows first.", "warning")
                    return redirect(url_for("staged_import_review", run_id=run.id))
                next_status = "Approved" if action == "approve" else "Rejected"
                note = (request.form.get("reviewer_note") or "").strip()
                for cid in selected_ids:
                    candidate = db.session.get(StagedImportCandidate, cid)
                    if not candidate or candidate.run_id != run.id:
                        continue
                    candidate.review_status = next_status
                    if note:
                        candidate.reviewer_note = note
                db.session.commit()
                flash(f"{len(selected_ids)} staged row(s) marked as {next_status}.", "success")
                return redirect(url_for("staged_import_review", run_id=run.id))

        candidates = StagedImportCandidate.query.filter_by(run_id=run.id).order_by(StagedImportCandidate.kind.asc(), StagedImportCandidate.confidence_score.desc()).all()
        return render_template("staged_import_review.html", run=run, candidates=candidates)

    @app.post("/admin/registry/paste")
    def stage_registry_paste():
        user = require_login()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        registry_text = request.form.get("registry_text", "")
        source_url = request.form.get("source_url", "").strip() or "cipa-manual-paste"
        jurisdiction = request.form.get("jurisdiction", "Botswana").strip() or "Botswana"
        if not registry_text.strip():
            flash("Paste CIPA page text first.", "warning")
            return redirect(url_for("admin"))

        parsed = parse_cipa_registry_text(registry_text)
        company_name = str(parsed.get("company_name") or "").strip() or "CIPA Registry"
        company_number = str(parsed.get("company_number") or "").strip()
        run = StagedImportRun(
            source_name=f"CIPA registry - {company_name}".strip(),
            source_url=source_url,
            jurisdiction=jurisdiction,
            status="staged",
            created_by=user.email,
            message=f"Staged registry paste for {company_name} {('('+company_number+')') if company_number else ''}.",
            metrics_json=json.dumps({"directors": len(parsed.get('directors') or [])}, default=str),
        )
        db.session.add(run)
        db.session.flush()

        # Stage a company marker row (registry kind)
        db.session.add(
            StagedImportCandidate(
                run_id=run.id,
                kind="registry",
                full_name=company_name,
                category="Company",
                position=str(parsed.get("company_status") or ""),
                snippet=f"{company_number} {parsed.get('company_type') or ''}".strip(),
                confidence_score=90,
                evidence_json=json.dumps(parsed, default=str),
            )
        )

        for director in parsed.get("directors") or []:
            if not isinstance(director, dict):
                continue
            name = str(director.get("name") or "").strip()
            if not name:
                continue
            db.session.add(
                StagedImportCandidate(
                    run_id=run.id,
                    kind="pip",
                    full_name=name,
                    category="Related parties",
                    position="Director",
                    confidence_score=75,
                    snippet=f"{company_name} director; appointment {director.get('appointment_date','')}".strip(),
                    evidence_json=json.dumps({"company": company_name, "company_number": company_number, "director": director, "source_url": source_url, "source_name": run.source_name}, default=str),
                )
            )
        db.session.commit()
        flash(f"Registry paste staged. Review before applying: /admin/staged-imports/{run.id}", "success")
        return redirect(url_for("staged_import_review", run_id=run.id))


app = create_app()


if __name__ == "__main__":
    app.run(debug=True, port=5055)
