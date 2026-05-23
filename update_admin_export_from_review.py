from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app import rejection_reason_for_name


EXPORT_PATH = Path(r"C:\Users\HP\Downloads\pep_pip_records_export_20260516.xlsx")
REVIEW_PATH = Path(r"C:\Users\HP\Downloads\pep_screening_template.xlsx")


def parse_notes(notes: str) -> dict[str, str]:
    parsed: dict[str, str] = {}
    for key in ["Category", "Position", "Jurisdiction", "Confidence", "Evidence"]:
        if key == "Evidence":
            match = re.search(r"Evidence:\s*(.+)", notes or "", flags=re.S)
        else:
            match = re.search(rf"{key}:\s*([^;]+)", notes or "")
        if match:
            parsed[key.lower()] = match.group(1).strip()
    return parsed


def row_map(headers: list[str], row_number: int) -> dict[str, int]:
    return {header: index + 1 for index, header in enumerate(headers)}


def main() -> None:
    backup_path = EXPORT_PATH.with_name(f"{EXPORT_PATH.stem}_backup_{datetime.now():%Y%m%d_%H%M%S}{EXPORT_PATH.suffix}")
    shutil.copy2(EXPORT_PATH, backup_path)

    export_wb = load_workbook(EXPORT_PATH)
    export_ws = export_wb["PEP Records"] if "PEP Records" in export_wb.sheetnames else export_wb.active
    headers = [str(export_ws.cell(1, column).value or "").strip() for column in range(1, export_ws.max_column + 1)]
    columns = row_map(headers, 1)

    existing_by_name: dict[str, int] = {}
    next_id = 0
    rejected = 0
    for row in range(2, export_ws.max_row + 1):
        full_name = str(export_ws.cell(row, columns["full_name"]).value or "").strip()
        if full_name:
            existing_by_name[full_name.lower()] = row
        try:
            next_id = max(next_id, int(export_ws.cell(row, columns["id"]).value or 0))
        except (TypeError, ValueError):
            pass
        if full_name and rejection_reason_for_name(full_name):
            export_ws.cell(row, columns["status"]).value = "Rejected / not a person"
            export_ws.cell(row, columns["adverse_media_status"]).value = "No adverse media"
            notes = str(export_ws.cell(row, columns["notes"]).value or "")
            if "Rejected during workbook review" not in notes:
                export_ws.cell(row, columns["notes"]).value = (notes + "\n\nRejected during workbook review: likely non-person/layout text.").strip()
            rejected += 1

    review_wb = load_workbook(REVIEW_PATH)
    review_ws = review_wb.active
    review_headers = [str(review_ws.cell(1, column).value or "").strip().lower() for column in range(1, review_ws.max_column + 1)]
    review_columns = {header: index + 1 for index, header in enumerate(review_headers)}

    created = 0
    updated = 0
    for row in range(2, review_ws.max_row + 1):
        name = str(review_ws.cell(row, review_columns["name"]).value or "").strip()
        if not name:
            continue
        reference = str(review_ws.cell(row, review_columns["reference"]).value or "").strip()
        notes = str(review_ws.cell(row, review_columns["notes"]).value or "").strip()
        parsed = parse_notes(notes)

        target_row = existing_by_name.get(name.lower())
        if not target_row:
            export_ws.append([None] * len(headers))
            target_row = export_ws.max_row
            next_id += 1
            export_ws.cell(target_row, columns["id"]).value = next_id
            export_ws.cell(target_row, columns["full_name"]).value = name
            existing_by_name[name.lower()] = target_row
            created += 1
        else:
            updated += 1

        export_ws.cell(target_row, columns["category"]).value = parsed.get("category") or export_ws.cell(target_row, columns["category"]).value or "Domestic PIP"
        export_ws.cell(target_row, columns["jurisdiction"]).value = parsed.get("jurisdiction") or export_ws.cell(target_row, columns["jurisdiction"]).value or "Botswana"
        export_ws.cell(target_row, columns["position"]).value = parsed.get("position") or export_ws.cell(target_row, columns["position"]).value
        export_ws.cell(target_row, columns["status"]).value = "Needs review"
        export_ws.cell(target_row, columns["adverse_media_status"]).value = export_ws.cell(target_row, columns["adverse_media_status"]).value or "Pending review"
        export_ws.cell(target_row, columns["source_type"]).value = "PDF"
        export_ws.cell(target_row, columns["source_name"]).value = "Database.pdf"
        export_ws.cell(target_row, columns["source_url"]).value = reference
        export_ws.cell(target_row, columns["source_excerpt"]).value = parsed.get("evidence", "")
        export_ws.cell(target_row, columns["reviewer_notes"]).value = "Imported from reviewed Database.pdf evidence; requires admin confirmation."
        export_ws.cell(target_row, columns["notes"]).value = notes

    export_ws.column_dimensions["B"].width = 32
    export_ws.column_dimensions["Q"].width = 90
    export_ws.column_dimensions["S"].width = 90
    export_wb.save(EXPORT_PATH)

    print(f"Backup: {backup_path}")
    print(f"Updated existing rows: {updated}")
    print(f"Created new rows: {created}")
    print(f"Marked non-person rows rejected: {rejected}")
    print(f"Updated workbook: {EXPORT_PATH}")


if __name__ == "__main__":
    main()
